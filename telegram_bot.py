import logging
import os
import pandas as pd
import io
import matplotlib
# Tkinter errors - use a non-interactive backend
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import telebot
from telebot import types
import numpy as np
import threading
import time
from data_processor import process_exam_data, prepare_excel_for_download, prepare_pdf_for_download, prepare_simplified_excel
from rasch_model import RaschModel
from utils import display_grade_distribution, GRADE_DESCRIPTIONS, calculate_statistics
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from bot_database import BotDatabase
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, PieChart

# Enable logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Keep track of user data
user_data = {}

# Foydalanuvchi ma'lumotlarini saqlash uchun dictionary
user_settings = {}

# Milliy sertifikat fanalari va ularning bo'limlari
SUBJECTS = {
    'matematika': {
        'name': 'Matematika',
        'sections': ['Algebra', 'Geometriya', 'Trigonometriya', 'Kalkulyus']
    },
    'ona_tili': {
        'name': 'Ona tili',
        'sections': ['Leksikologiya', 'Fonetika', 'Morfologiya', 'Sintaksis', 'Adabiyot']
    },
    'fizika': {
        'name': 'Fizika',
        'sections': ['Mexanika', 'Elektromagnetizm', 'Optika', 'Atom fizikasi']
    },
    'kimyo': {
        'name': 'Kimyo',
        'sections': ['Anorganik kimyo', 'Organik kimyo', 'Analitik kimyo']
    },
    'biologiya': {
        'name': 'Biologiya',
        'sections': ['Botanika', 'Zoologiya', 'Anatomiya', 'Genetika']
    },
    'tarix': {
        'name': 'Tarix',
        'sections': ['O\'zbekiston tarixi', 'Jahon tarixi', 'Madaniyat tarixi']
    },
    'geografiya': {
        'name': 'Geografiya',
        'sections': ['Fizik geografiya', 'Iqtisodiy geografiya', 'O\'zbekiston geografiyasi']
    },
    'ingliz_tili': {
        'name': 'Ingliz tili',
        'sections': ['Grammar', 'Vocabulary', 'Reading', 'Writing', 'Speaking']
    }
}

# Placeholder matnlar ro'yxati - turli xil jarayon xabarlari
PROCESS_PLACEHOLDER_MESSAGES = [
    "⏳ Hisoblanmoqda... biroz kuting!",
    "⌛ Iltimos, kutib turing...",
    "📊 Natijalar tayyorlanmoqda...",
    "🧮 DTM foizlari hisoblanmoqda...",
    "📈 Rasch model qo'llanilmoqda...",
    "🔍 Savollar qiyinligi aniqlanmoqda...",
    "📉 Grafiklar tayyorlanmoqda...",
    "⚙️ Tahlil jarayoni davom etmoqda...",
    "🧠 Ma'lumotlar tahlil qilinmoqda...",
    "🔢 Baholar hisoblanmoqda...",
    "📋 Natijalar formatlanmoqda...",
    "📝 Hisobot tayyorlanmoqda..."
]

# Tasodifiy placeholder tanlash va avtomatik almashtirib turish uchun funksiya
def get_random_placeholder():
    """Placeholder matnlar ro'yxatidan tasodifiy birini qaytaradi"""
    import random
    return random.choice(PROCESS_PLACEHOLDER_MESSAGES)

# Help message
HELP_MESSAGE = """
🚀 Rasch Counter Bot - Qo'llanma

Bot haqida: 
📊 Bu bot test natijalarini Rasch model orqali tahlil qilib, talabalarning bilim darajasini aniq baholashga yordam beradi.

Asosiy buyruqlar:
🔹 /start - Botni ishga tushirish
🔹 /help - Yordam olish
🔹 /ball - Ikki Excel fayldan o'rtacha ball hisoblash
🔹 /cancel - Joriy jarayonni bekor qilish

Excel fayl qanday bo'lishi kerak:
🔸 .xlsx yoki .xls formatda bo'lishi kerak
🔸 Birinchi ustunda talabaning ismi yoki ID raqami bo'lishi kerak
🔸 Qolgan ustunlarda savollar javoblari: 1 (to'g'ri) yoki 0 (noto'g'ri)
🔸 Har bir qator - bir talaba ma'lumotlari
🔸 Har bir ustun - bir savol javobi

Bot sizga quyidagi ma'lumotlarni beradi:
📈 Umumiy statistika (o'rtacha ball, o'zlashtirish foizi)
📊 Talabalarning baholar bo'yicha taqsimlanishi
📋 Qaysi savollar qiyin, qaysilari oson ekanligi
🔍 Test savollarining sifat ko'rsatkichlari

Natijalarni yuklab olish:
💾 Excel formatda - barcha batafsil ma'lumotlar bilan
📑 PDF formatda - chiroyli formatda taqdim etilgan natijalar

Botdan foydalanish:
1️⃣ Excel faylni yuklang
2️⃣ Bot tahlilni avtomatik ravishda boshlaydi
3️⃣ Natijalar tayyor bo'lgach, tugmalar orqali kerakli ma'lumotlarni ko'ring

Savollar va takliflar uchun:
📬 Barcha savol va takliflaringizni ushbu guruhga yuboring: t.me/rasch_counter
"""

def migrate_database():
    """Migrate database from root directory to .data directory if needed"""
    old_db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'bot_data.db')
    
    # Check if old database exists in root directory
    if os.path.exists(old_db_path):
        # Create .data directory if it doesn't exist
        data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.data')
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)
        
        new_db_path = os.path.join(data_dir, 'bot_data.db')
        
        # Only copy if the target doesn't exist
        if not os.path.exists(new_db_path):
            import shutil
            shutil.copy2(old_db_path, new_db_path)
            print(f"Database migrated from {old_db_path} to {new_db_path}")
    
    return True

def main():
    """Start the bot."""
    # Get the telegram token
    token = "8061680273:AAGgJhdvWUwnP1JJcgk_AksRxSJSs_VbvLk"
    if not token:
        logger.error("TELEGRAM_TOKEN not found in environment variables.")
        print("Please set the TELEGRAM_TOKEN environment variable.")
        return
    
    # Migrate database if needed
    migrate_database()
    
    # Create bot instance
    bot = telebot.TeleBot(token)
    
    # Initialize database
    db = BotDatabase()
    
    # Admin command handler - only accessible by specific admin user ID
    @bot.message_handler(commands=['adminos'])
    def admin_command(message):
        # Check if the user is authorized admin (specific Telegram ID)
        if message.from_user.id != 7537966029:
            bot.send_message(
                message.chat.id,
                "⛔ Kechirasiz, bu buyruq faqat adminlar uchun.\n"
                "Sorry, this command is only for admins."
            )
            return
        
        # Get overall statistics for admin panel
        total_stats = db.get_user_stats()
        all_users = db.get_all_users()
        
        # Create admin keyboard with only broadcast option
        markup = types.InlineKeyboardMarkup(row_width=1)
        btn_reklama = types.InlineKeyboardButton('Reklama yuborish', callback_data='admin_broadcast')
        markup.add(btn_reklama)
        
        # Create message with basic statistics
        admin_message = f"🔐 *Admin paneli*\n\n"
        admin_message += f"📊 *Statistika*:\n"
        admin_message += f"• Foydalanuvchilar: {len(all_users)} ta\n"
        admin_message += f"• Tekshirilgan testlar: {total_stats.get('exam_count', 0)} ta\n"
        admin_message += f"• Tekshirilgan talabalar: {total_stats.get('total_students', 0)} ta\n\n"
        admin_message += f"Quyidagi buyruqdan foydalaning:"
        
        bot.send_message(
            message.chat.id,
            admin_message,
            reply_markup=markup,
            parse_mode='Markdown'
        )
    
    # Start command handler
    @bot.message_handler(commands=['start'])
    def send_welcome(message):
        """Start command handler"""
        user_name = message.from_user.first_name or "foydalanuvchi"
        welcome_text = f"""
👋 Assalomu alaykum, {user_name} !

🎓 Rasch Counter Botga xush kelibsiz!

📝 Excel faylni yuboring va natijalarni oling
"""
        bot.send_message(message.chat.id, welcome_text, parse_mode='Markdown')
        
        # Alohida davomiy xabar
        follow_up_text = "📊 Excel faylni yuborishingiz mumkin"
        bot.send_message(message.chat.id, follow_up_text)
    
    # Help command handler
    @bot.message_handler(commands=['help'])
    def handle_help(message):
        """Yordam ma'lumotlari"""
        help_text = """
📚 *Bot yordam ma'lumotlari*

🎯 *Asosiy funksiyalar:*
• Rasch modeli asosida test natijalarini tahlil qilish
• Fan bo'limlari bo'yicha alohida salohiyat baholash
• Professional hisobotlar yaratish

⚙️ *Sozlash jarayoni:*
1. `/setting` - Fan tanlash
2. Har bir bo'lim uchun savol raqamlarini kiriting
3. Natijalar matritsasini yuboring

📊 *Hisobot turlari:*
• 💾 Excel formatda yuklash - To'liq ma'lumotlar
• 📑 PDF formatda yuklash - Chiroyli format
• 📝 Nazorat Ballari - Oddiy natijalar

❓ Savollaringiz bo'lsa, administrator bilan bog'laning.
"""
        bot.reply_to(message, help_text, parse_mode='Markdown')
    
    # Ball command handler
    @bot.message_handler(commands=['ball'])
    def ball_command(message):
        """Start process to calculate average scores from two files."""
        # Send instructions
        instruction_text = """
        <b>O'rtacha ball hisoblash uchun:</b>
        
        1. Ikkita Excel faylni ketma-ket yuboring
        2. Har bir faylda "Talaba" va "Ball" ustunlari bo'lishi kerak
        3. Bot ikkala fayl asosida o'rtacha ballni hisoblab, yangi fayl qaytaradi
        
        Iltimos, birinchi Excel faylni yuboring.
        
        <i>Jarayonni bekor qilish uchun /cancel yoki /start buyrug'ini yuboring.</i>
        """
        
        # Save user state to wait for first file
        user_data[message.chat.id] = {
            'waiting_for_balls': 'first_file',
            'first_file': None,
            'second_file': None
        }
        
        bot.send_message(chat_id=message.chat.id, text=instruction_text, parse_mode='HTML')
        
    # Fill command handler - disabled/removed as requested
            
    # Cancel command handler
    @bot.message_handler(commands=['cancel'])
    def cancel_command(message):
        user_id = message.from_user.id
        
        if user_id in user_data:
            # Check for broadcast mode
            if 'waiting_for_broadcast' in user_data[user_id]:
                broadcast_type = user_data[user_id]['waiting_for_broadcast']
                # Reset the user state
                user_data[user_id] = {}
                
                # Any broadcast type can be cancelled with a standardized message
                bot.send_message(
                    message.chat.id,
                    "❌ Xabar yuborish bekor qilindi."
                )
                return
            
            # Check for ball mode
            if 'waiting_for_balls' in user_data[user_id]:
                # Reset the user state
                user_data[user_id] = {}
                
                bot.send_message(
                    message.chat.id,
                    "❌ Ball hisoblash jarayoni bekor qilindi.\n\n"
                    "Test tahlili rejimiga qaytdingiz. Excel faylni yuborishingiz mumkin."
                )
                return
        
        # If no active process found
        bot.send_message(
            message.chat.id,
            "⚠️ Bekor qilinadigan jarayon yo'q."
        )
    
    # Text message handler for broadcast
    @bot.message_handler(func=lambda message: message.from_user.id in user_data 
                                        and 'waiting_for_broadcast' in user_data[message.from_user.id]
                                        and (user_data[message.from_user.id]['waiting_for_broadcast'] == 'text' 
                                             or user_data[message.from_user.id]['waiting_for_broadcast'] == 'any'))
    def handle_broadcast_text(message):
        # Only allow the admin to broadcast
        if message.from_user.id != 7537966029:
            return
            
        broadcast_text = message.text
        user_id = message.from_user.id
        
        # Reset the admin's state
        user_data[user_id] = {}
        
        # Get all users from database
        all_users = db.get_all_users()
        
        # Send status message
        status_message = bot.send_message(
            message.chat.id,
            f"⏳ {len(all_users)} ta foydalanuvchiga xabar yuborilmoqda..."
        )
        
        # Track successful/failed sends
        success_count = 0
        failed_count = 0
        
        # Send the message to all users - without any prefix
        for user in all_users:
            try:
                bot.send_message(
                    user['user_id'],
                    broadcast_text,
                    parse_mode='Markdown'
                )
                success_count += 1
                
                # Update status every 10 users
                if success_count % 10 == 0:
                    bot.edit_message_text(
                        chat_id=message.chat.id,
                        message_id=status_message.message_id,
                        text=f"⏳ {success_count}/{len(all_users)} ta foydalanuvchiga xabar yuborildi..."
                    )
                    
                # Avoid hitting Telegram's rate limits
                time.sleep(0.1)
                
            except Exception as e:
                failed_count += 1
                logger.error(f"Failed to send broadcast to user {user['user_id']}: {str(e)}")
        
        # Send completion message
        bot.edit_message_text(
            chat_id=message.chat.id,
            message_id=status_message.message_id,
            text=f"✅ Xabar yuborish yakunlandi!\n\n"
                 f"• Yuborildi: {success_count} ta\n"
                 f"• Xatolik: {failed_count} ta"
        )
    
    # Image handler for broadcast
    @bot.message_handler(content_types=['photo'], func=lambda message: message.from_user.id in user_data 
                                          and 'waiting_for_broadcast' in user_data[message.from_user.id]
                                          and (user_data[message.from_user.id]['waiting_for_broadcast'] == 'image'
                                               or user_data[message.from_user.id]['waiting_for_broadcast'] == 'any'))
    def handle_broadcast_image(message):
        # Only allow the admin to broadcast
        if message.from_user.id != 7537966029:
            return
            
        user_id = message.from_user.id
        
        # Get caption if any
        caption = message.caption or ""
        
        # Get the photo file_id (highest quality version)
        photo_id = message.photo[-1].file_id
        
        # Reset the admin's state
        user_data[user_id] = {}
        
        # Get all users from database
        all_users = db.get_all_users()
        
        # Send status message
        status_message = bot.send_message(
            message.chat.id,
            f"⏳ {len(all_users)} ta foydalanuvchiga rasm yuborilmoqda..."
        )
        
        # Track successful/failed sends
        success_count = 0
        failed_count = 0
        
        # Send the image to all users
        for user in all_users:
            try:
                bot.send_photo(
                    user['user_id'],
                    photo=photo_id,
                    caption=caption,
                    parse_mode='Markdown'
                )
                success_count += 1
                
                # Update status every 10 users
                if success_count % 10 == 0:
                    bot.edit_message_text(
                        chat_id=message.chat.id,
                        message_id=status_message.message_id,
                        text=f"⏳ {success_count}/{len(all_users)} ta foydalanuvchiga rasm yuborildi..."
                    )
                    
                # Avoid hitting Telegram's rate limits
                time.sleep(0.1)
                
            except Exception as e:
                failed_count += 1
                logger.error(f"Failed to send broadcast image to user {user['user_id']}: {str(e)}")
        
        # Send completion message
        bot.edit_message_text(
            chat_id=message.chat.id,
            message_id=status_message.message_id,
            text=f"✅ Rasm yuborish yakunlandi!\n\n"
                 f"• Yuborildi: {success_count} ta\n"
                 f"• Xatolik: {failed_count} ta"
        )
        
    # Video handler for broadcast
    @bot.message_handler(content_types=['video'], func=lambda message: message.from_user.id in user_data 
                                          and 'waiting_for_broadcast' in user_data[message.from_user.id]
                                          and (user_data[message.from_user.id]['waiting_for_broadcast'] == 'video'
                                               or user_data[message.from_user.id]['waiting_for_broadcast'] == 'any'))
    def handle_broadcast_video(message):
        # Only allow the admin to broadcast
        if message.from_user.id != 7537966029:
            return
            
        user_id = message.from_user.id
        
        # Get caption if any
        caption = message.caption or ""
        
        # Get the video file_id
        video_id = message.video.file_id
        
        # Reset the admin's state
        user_data[user_id] = {}
        
        # Get all users from database
        all_users = db.get_all_users()
        
        # Send status message
        status_message = bot.send_message(
            message.chat.id,
            f"⏳ {len(all_users)} ta foydalanuvchiga video yuborilmoqda..."
        )
        
        # Track successful/failed sends
        success_count = 0
        failed_count = 0
        
        # Send the video to all users
        for user in all_users:
            try:
                bot.send_video(
                    user['user_id'],
                    video=video_id,
                    caption=caption,
                    parse_mode='Markdown'
                )
                success_count += 1
                
                # Update status every 5 users to avoid flooding
                if success_count % 5 == 0:
                    bot.edit_message_text(
                        chat_id=message.chat.id,
                        message_id=status_message.message_id,
                        text=f"⏳ {success_count}/{len(all_users)} ta foydalanuvchiga video yuborildi..."
                    )
                    
                # Avoid hitting Telegram's rate limits (longer delay for videos)
                time.sleep(0.2)
                
            except Exception as e:
                failed_count += 1
                logger.error(f"Failed to send broadcast video to user {user['user_id']}: {str(e)}")
        
        # Send completion message
        bot.edit_message_text(
            chat_id=message.chat.id,
            message_id=status_message.message_id,
            text=f"✅ Video yuborish yakunlandi!\n\n"
                 f"• Yuborildi: {success_count} ta\n"
                 f"• Xatolik: {failed_count} ta"
        )
        
    # Sticker handler for broadcast
    @bot.message_handler(content_types=['sticker'], func=lambda message: message.from_user.id in user_data 
                                          and 'waiting_for_broadcast' in user_data[message.from_user.id]
                                          and (user_data[message.from_user.id]['waiting_for_broadcast'] == 'sticker'
                                               or user_data[message.from_user.id]['waiting_for_broadcast'] == 'any'))
    def handle_broadcast_sticker(message):
        # Only allow the admin to broadcast
        if message.from_user.id != 7537966029:
            return
            
        user_id = message.from_user.id
        
        # Get the sticker file_id
        sticker_id = message.sticker.file_id
        
        # Reset the admin's state
        user_data[user_id] = {}
        
        # Get all users from database
        all_users = db.get_all_users()
        
        # Send status message
        status_message = bot.send_message(
            message.chat.id,
            f"⏳ {len(all_users)} ta foydalanuvchiga stiker yuborilmoqda..."
        )
        
        # Track successful/failed sends
        success_count = 0
        failed_count = 0
        
        # Send the sticker to all users
        for user in all_users:
            try:
                bot.send_sticker(
                    user['user_id'],
                    sticker=sticker_id
                )
                success_count += 1
                
                # Update status every 10 users
                if success_count % 10 == 0:
                    bot.edit_message_text(
                        chat_id=message.chat.id,
                        message_id=status_message.message_id,
                        text=f"⏳ {success_count}/{len(all_users)} ta foydalanuvchiga stiker yuborildi..."
                    )
                    
                # Avoid hitting Telegram's rate limits
                time.sleep(0.1)
                
            except Exception as e:
                failed_count += 1
                logger.error(f"Failed to send broadcast sticker to user {user['user_id']}: {str(e)}")
        
        # Send completion message
        bot.edit_message_text(
            chat_id=message.chat.id,
            message_id=status_message.message_id,
            text=f"✅ Stiker yuborish yakunlandi!\n\n"
                 f"• Yuborildi: {success_count} ta\n"
                 f"• Xatolik: {failed_count} ta"
        )
    
    # File handler
    @bot.message_handler(content_types=['document'])
    def handle_document(message):
        user_id = message.from_user.id
        
        # Get file info
        file_info = message.document
            
        # Regular Excel file handling (not in fill mode)
        if not file_info.file_name.endswith(('.xlsx', '.xls')):
            bot.send_message(
                message.chat.id, 
                "⚠️ Kechirasiz, bu fayl formati qo'llab-quvvatlanmaydi.\n\n"
                "ℹ️ Faqat Excel fayllarini (.xlsx, .xls) yuborishingiz mumkin.\n\n"
                "🔄 Iltimos, to'g'ri formatdagi faylni yuboring."
            )
            return
        
        # Check if user is in /ball command mode
        if message.chat.id in user_data and 'waiting_for_balls' in user_data[message.chat.id]:
            handle_ball_file(message, file_info)
            return
        
        # Fayl yuklab olinganini bildirish
        excel_message = bot.send_message(
            message.chat.id, 
            "📊 Excel fayl qabul qilindi!\n"
            "📂 Fayl yuklab olindi."
        )
        
        # Faylga emoji bilan reaksiya ko'rsatish - test botda tasdiqlangan usul
        try:
            import requests
            # Telegram API URL
            base_url = f"https://api.telegram.org/bot{bot.token}"
            
            # 1-usul (test botda yaxshi ishladi)
            reaction_data = {
                'chat_id': message.chat.id,
                'message_id': message.message_id,
                'reaction': [{'type': 'emoji', 'emoji': '👌'}]
            }
            response = requests.post(f"{base_url}/setMessageReaction", json=reaction_data)
            
            # Agar xatolik bo'lsa, log'ga yozamiz
            if response.status_code != 200:
                logger.error(f"Reaksiya qo'shishda xatolik: {response.status_code} - {response.text}")
                # Agar reaksiya ishlamasa, oddiy emojili javob beramiz
                bot.send_message(
                    message.chat.id,
                    "👌",
                    reply_to_message_id=message.message_id
                )
        except Exception as e:
            logger.error(f"Reaksiya qo'shishda xatolik: {str(e)}")
            # Xatolik bo'lsa, oddiy emojili javob beramiz
            bot.send_message(
                message.chat.id,
                "👌",
                reply_to_message_id=message.message_id
            )
        
        # Ishlov berish jarayoni boshlangani haqida xabar - har safar tasodifiy matn bilan
        process_message = bot.send_message(
            message.chat.id, 
            f"⏳ {get_random_placeholder()}",
            reply_markup=None
        )
        
        try:
            # Faylni ishlov berish jarayoni
            import time
            
            # Fayl yuklanmoqda...
            
            # Faylni yuklab olish
            file_id = file_info.file_id
            file_info = bot.get_file(file_id)
            downloaded_file = bot.download_file(file_info.file_path)
            file_bytes = io.BytesIO(downloaded_file)
            
            # Ma'lumotlarni o'qish
            df = pd.read_excel(file_bytes)
            
            # Process the data with improved information
            results_df, ability_estimates, grade_counts, data_df, beta_values, rasch_model_obj = process_exam_data(df)
            
            # Track user activity in database
            db.add_user(
                user_id=message.from_user.id,
                first_name=message.from_user.first_name,
                last_name=message.from_user.last_name or "",
                username=message.from_user.username or ""
            )
            
            # Log file processing with statistics
            db.log_file_processing(
                user_id=message.from_user.id,
                action_type="process_exam",
                num_students=len(results_df),
                num_questions=len(data_df.columns) - 1  # Subtract 1 for student name column
            )
            
            # Prepare Excel data for download
            excel_data = prepare_excel_for_download(results_df)
            
            # Store results for this user, including item difficulties and original data
            user_data[user_id] = {
                'results_df': results_df,
                'ability_estimates': ability_estimates,
                'grade_counts': grade_counts,
                'excel_data': excel_data,
                'data_df': data_df,        # Original data with student responses
                'beta_values': beta_values, # Item difficulty parameters from Rasch model
                'original_df': df,          # Original unprocessed data
                'rasch_model': rasch_model_obj  # Rasch model object with fit statistics
            }
            
            # We no longer need to send the comparison file automatically
            # The results are sufficient if they are successfully processed
            
            # Create keyboard with buttons - yaxshilangan tuzilish
            markup = types.InlineKeyboardMarkup(row_width=2)
            
            btn_statistics = types.InlineKeyboardButton('📊 Statistika', callback_data='statistics')
            btn_excel = types.InlineKeyboardButton('💾 Excel formatda yuklash', callback_data='download_excel')
            btn_pdf = types.InlineKeyboardButton('📑 PDF formatda yuklash', callback_data='download_pdf')
            btn_simple_excel = types.InlineKeyboardButton('📝 Nazorat Ballari', callback_data='download_simple_excel')
            
            markup.add(btn_statistics)
            markup.add(btn_excel, btn_pdf)
            markup.add(btn_simple_excel)
            
            # Emojilar bilan ma'noli javob
            # A+/A baholar soni uchun
            top_grades_count = grade_counts.get('A+', 0) + grade_counts.get('A', 0)
            # B+/B/C+/C baholar soni uchun
            passing_grades_count = top_grades_count + grade_counts.get('B+', 0) + grade_counts.get('B', 0) + grade_counts.get('C+', 0) + grade_counts.get('C', 0)
            # Sertifikat ololmaganlar soni (NC - No Certificate)
            failing_count = grade_counts.get('NC', 0)
            
            # Umumiy o'tish foizini hisoblash
            pass_rate = (passing_grades_count / len(results_df) * 100) if len(results_df) > 0 else 0
            
            # Natijani chiqarish
            
            # Natija tayyorligi haqida xabar
            try:
                bot.edit_message_text(
                    chat_id=message.chat.id,
                    message_id=process_message.message_id,
                    text="✅ Tahlil muvaffaqiyatli yakunlandi!"
                )
                time.sleep(1)  # Natija chiqganini ko'rish uchun 1 soniya kutish
                
                # Endi eski xabarni olib tashlaymiz
                bot.delete_message(chat_id=message.chat.id, message_id=process_message.message_id)
            except Exception as e:
                print(f"Xabarni yangilash/o'chirishda xatolik: {str(e)}")
                # Xatolikni e'tiborsiz qoldiramiz
                
            # Natijalar haqida qisqa ma'lumot
            # Nolga bo'linish xatosidan himoya
            total_students = len(results_df)
            top_grade_percent = (top_grades_count/total_students*100) if total_students > 0 else 0
            failing_percent = (failing_count/total_students*100) if total_students > 0 else 0
            
            success_message = (
                f"✅ Tahlil yakunlandi!\n\n"
                f"📊 Natijalar xulosasi:\n"
                f"👨‍🎓 Jami: {total_students} talaba\n"
                f"🏆 A+/A: {top_grades_count} ta ({top_grade_percent:.1f}%)\n"
                f"✅ O'tish: {passing_grades_count} ta ({pass_rate:.1f}%)\n"
                f"❌ O'tmagan: {failing_count} ta ({failing_percent:.1f}%)\n\n"
                f"📈 Quyidagi tugmalardan birini tanlang 👇"
            )
            
            bot.send_message(
                message.chat.id,
                success_message,
                reply_markup=markup
            )
            
        except Exception as e:
            # Xatolik yuz bergani haqida xabar berish
            
            # Xatolik haqida log yozish
            logger.error(f"Error processing file: {str(e)}")
            
            # Avval hisoblanmoqda... xabarini yangilaymiz
            try:
                bot.edit_message_text(
                    chat_id=message.chat.id,
                    message_id=process_message.message_id,
                    text="❌ Xatolik yuz berdi! Fayl bilan muammo bor."
                )
                time.sleep(1)  # Xatolik xabarini ko'rish uchun 1 soniya kutish
                
                # Endi eski xabarni olib tashlaymiz
                bot.delete_message(chat_id=message.chat.id, message_id=process_message.message_id)
            except Exception as msg_error:
                print(f"Xabarni yangilash/o'chirishda xatolik: {str(msg_error)}")
                # Xatolikni e'tiborsiz qoldiramiz
            
            # Foydalanuvchiga xatolik haqida batafsil ma'lumot beramiz
            bot.send_message(
                message.chat.id,
                f"❌ Xatolik yuz berdi!\n\n"
                f"⚠️ Muammo tavsifi: {str(e)}\n\n"
                f"📋 Excel fayl quyidagi talablarga javob berishi kerak:\n"
                f"1️⃣ Birinchi ustunda talaba ID/ismi bo'lishi kerak\n"
                f"2️⃣ Har bir savol 1 (to'g'ri) yoki 0 (noto'g'ri) qiymatlardan iborat bo'lishi kerak\n"
                f"3️⃣ Fayl tuzilishi: har bir qator = bir talaba, har bir ustun = bir savol\n\n"
                f"🔄 Iltimos, faylni tekshirib, qayta yuboring."
            )
    
    # Callback handler for inline buttons
    @bot.callback_query_handler(func=lambda call: True)
    def callback_query(call):
        try:
            # Extract callback data
            callback_data = call.data
            
            # Handle different callback types
            if callback_data.startswith('download_'):
                handle_download_callback(call)
            elif callback_data == 'statistics':
                handle_statistics_callback(call)
            elif callback_data == 'all_results':
                handle_all_results_callback(call)
            else:
                # Default handling
                handle_default_callback(call)
                
        except Exception as e:
            logger.error(f"Callback xatosi: {e}")
            bot.answer_callback_query(call.id, "Xatolik yuz berdi. Iltimos, qaytadan urinib ko'ring.")
    
    def handle_download_callback(call):
        """Download callback handling"""
        try:
            user_id = call.from_user.id
            if user_id not in user_data or 'results_df' not in user_data[user_id]:
                bot.answer_callback_query(call.id, "Ma'lumotlar topilmadi. Iltimos, faylni qaytadan yuklang.")
                return
            
            results_df = user_data[user_id]['results_df']
            
            if call.data == 'download_excel':
                # Excel download
                excel_buffer = prepare_excel_for_download(results_df)
                bot.send_document(call.message.chat.id, 
                                ('natijalar.xlsx', excel_buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
                bot.answer_callback_query(call.id, "Excel fayl yuklandi!")
                
            elif call.data == 'download_pdf':
                # PDF download
                pdf_buffer = prepare_pdf_for_download(results_df)
                bot.send_document(call.message.chat.id, 
                                ('natijalar.pdf', pdf_buffer.getvalue(), 'application/pdf'))
                bot.answer_callback_query(call.id, "PDF fayl yuklandi!")
                
        except Exception as e:
            logger.error(f"Download callback xatosi: {e}")
            bot.answer_callback_query(call.id, "Yuklashda xatolik yuz berdi.")
    

    
    def handle_statistics_callback(call):
        """Yagona statistika callback - barcha ma'lumotlar va grafiklar"""
        try:
            user_id = call.from_user.id
            if user_id not in user_data:
                bot.answer_callback_query(call.id, "Ma'lumotlar topilmadi. Iltimos, faylni qaytadan yuklang.")
                return
            
            results_df = user_data[user_id].get('results_df')
            rasch_model = user_data[user_id].get('rasch_model')
            
            if results_df is None:
                bot.answer_callback_query(call.id, "Natijalar topilmadi.")
                return
            
            # 1. Statistika matni yaratish
            stats_text = create_comprehensive_statistics(results_df, rasch_model)
            
            # 2. Wright map yaratish
            wright_map_buffer = None
            if rasch_model:
                fig = rasch_model.create_wright_map()
                if fig:
                    wright_map_buffer = io.BytesIO()
                    fig.savefig(wright_map_buffer, format='png', dpi=300, bbox_inches='tight')
                    wright_map_buffer.seek(0)
                    plt.close(fig)
            
            # 3. Fit statistikalar grafigi
            fit_stats_buffer = None
            if rasch_model and hasattr(rasch_model, 'infit_stats'):
                from utils import create_fit_statistics_plot
                fig = create_fit_statistics_plot(rasch_model)
                if fig:
                    fit_stats_buffer = io.BytesIO()
                    fig.savefig(fit_stats_buffer, format='png', dpi=300, bbox_inches='tight')
                    fit_stats_buffer.seek(0)
                    plt.close(fig)
            
            # 4. Statistika Excel fayli yaratish
            stats_excel_buffer = create_statistics_excel(results_df, rasch_model)
            
            # 5. Natijalarni yuborish
            # Asosiy statistika matni + inline tugmalar
            markup = types.InlineKeyboardMarkup(row_width=2)
            btn_excel = types.InlineKeyboardButton('💾 Excel formatda yuklash', callback_data='download_excel')
            btn_pdf = types.InlineKeyboardButton('📑 PDF formatda yuklash', callback_data='download_pdf')
            btn_simple_excel = types.InlineKeyboardButton('📝 Nazorat Ballari', callback_data='download_simple_excel')
            
            markup.add(btn_excel, btn_pdf)
            markup.add(btn_simple_excel)
            
            bot.edit_message_text(stats_text, call.message.chat.id, call.message.message_id, 
                                parse_mode='Markdown', reply_markup=markup)
            
            # Wright map
            if wright_map_buffer:
                bot.send_photo(call.message.chat.id, wright_map_buffer,
                              caption="🗺️ **Wright Map (Item-Person Xaritasi)**\n\nTalabalar qobiliyati va savollar qiyinligi taqsimoti")
            
            # Fit statistikalar grafigi
            if fit_stats_buffer:
                bot.send_photo(call.message.chat.id, fit_stats_buffer,
                              caption="📈 **Fit Statistikalar**\n\nInfit va Outfit ko'rsatkichlari")
            
            # Statistika Excel fayli
            if stats_excel_buffer:
                bot.send_document(call.message.chat.id,
                                ('statistika.xlsx', stats_excel_buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                                caption="📊 **To'liq Statistika Hisoboti**\n\nBarcha statistik ma'lumotlar va jadvallar")
            
            bot.answer_callback_query(call.id, "📊 Statistika tayyorlandi!")
            
        except Exception as e:
            logger.error(f"Statistics callback xatosi: {e}")
            bot.answer_callback_query(call.id, "Statistika yaratishda xatolik yuz berdi.")
    
    def handle_all_results_callback(call):
        """Barcha natijalar callback - eski funksiya"""
        try:
            user_id = call.from_user.id
            if user_id not in user_data:
                bot.answer_callback_query(call.id, "Ma'lumotlar topilmadi.")
                return
            
            results_df = user_data[user_id].get('results_df')
            if results_df is None:
                bot.answer_callback_query(call.id, "Natijalar topilmadi.")
                return
            
            # Oddiy natijalar ko'rsatish
            stats_text = calculate_statistics(results_df)
            bot.edit_message_text(stats_text, call.message.chat.id, call.message.message_id)
            bot.answer_callback_query(call.id)
            
        except Exception as e:
            logger.error(f"All results callback xatosi: {e}")
            bot.answer_callback_query(call.id, "Natijalarni ko'rishda xatolik yuz berdi.")
    
    def handle_default_callback(call):
        """Default callback handling"""
        bot.answer_callback_query(call.id, "Bu funksiya hali ishlab chiqilmoqda.")
        user_id = call.from_user.id
        bot.answer_callback_query(call.id)
        
        # Handle admin callbacks
        if call.data.startswith("admin_"):
            # Verify this is the admin
            if user_id != 7537966029:
                bot.send_message(
                    call.message.chat.id,
                    "⛔ Kechirasiz, bu amalni faqat adminlar bajarishi mumkin."
                )
                return
                

                
            # Handle broadcast message
            elif call.data == "admin_broadcast":
                # Set state to wait for any broadcast message
                user_data[user_id] = {'waiting_for_broadcast': 'any'}
                
                # Ask admin for broadcast message
                bot.send_message(
                    call.message.chat.id,
                    "📣 *Xabar yuborish*\n\n"
                    "Barcha foydalanuvchilarga yubormoqchi bo'lgan xabarni yuboring.\n"
                    "Bu xabar barcha foydalanuvchilarga yuboriladi.\n\n"
                    "Siz yuborgan xabar turi avtomatik ravishda aniqlanadi:\n"
                    "- Matn xabar\n"
                    "- Rasm\n"
                    "- Video\n"
                    "- Stiker\n\n"
                    "Bekor qilish uchun /cancel buyrug'ini yuboring.",
                    parse_mode='Markdown'
                )
                return
                

                
            # Handle back button from admin submenus
            elif call.data == "admin_back":
                # Return to main admin panel with statistics
                # Get overall statistics for admin panel
                total_stats = db.get_user_stats()
                all_users = db.get_all_users()
                
                # Create admin keyboard with only broadcast option
                markup = types.InlineKeyboardMarkup(row_width=1)
                btn_reklama = types.InlineKeyboardButton('Reklama yuborish', callback_data='admin_broadcast')
                markup.add(btn_reklama)
                
                # Create message with basic statistics
                admin_message = f"🔐 *Admin paneli*\n\n"
                admin_message += f"📊 *Statistika*:\n"
                admin_message += f"• Foydalanuvchilar: {len(all_users)} ta\n"
                admin_message += f"• Tekshirilgan testlar: {total_stats.get('exam_count', 0)} ta\n"
                admin_message += f"• Tekshirilgan talabalar: {total_stats.get('total_students', 0)} ta\n\n"
                admin_message += f"Quyidagi buyruqdan foydalaning:"
                
                bot.edit_message_text(
                    chat_id=call.message.chat.id,
                    message_id=call.message.message_id,
                    text=admin_message,
                    reply_markup=markup,
                    parse_mode='Markdown'
                )
                return
                
            return
        
        # Handle regular user callbacks
        if user_id not in user_data:
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text="⚠️ Ma'lumotlar topilmadi!\n\n"
                     "Siz avval yuborgan Excel fayllarning ma'lumotlari saqlanmagan ko'rinadi.\n\n"
                     "📤 Iltimos, qaytadan Excel faylini yuboring."
            )
            return
        
        user_info = user_data[user_id]
        
        # Check if we're waiting for broadcast message
        if 'waiting_for_broadcast' in user_info and user_info['waiting_for_broadcast']:
            return
            
        results_df = user_info['results_df']
        grade_counts = user_info['grade_counts']
        
        if call.data == "back_to_menu":
            # Return to the main menu with original buttons
            markup = types.InlineKeyboardMarkup(row_width=2)
            
            btn_all_results = types.InlineKeyboardButton('📊 Barcha Natijalar va Grafiklar', callback_data='all_results')
            btn_excel = types.InlineKeyboardButton('💾 Excel formatda yuklash', callback_data='download_excel')
            btn_pdf = types.InlineKeyboardButton('📑 PDF formatda yuklash', callback_data='download_pdf')
            btn_simple_excel = types.InlineKeyboardButton('📝 Nazorat Ballari', callback_data='download_simple_excel')
            
            markup.add(btn_all_results)
            markup.add(btn_excel, btn_pdf)
            markup.add(btn_simple_excel)
            
            # A+/A baholar soni uchun
            top_grades_count = grade_counts.get('A+', 0) + grade_counts.get('A', 0)
            # B+/B/C+/C baholar soni uchun
            passing_grades_count = top_grades_count + grade_counts.get('B+', 0) + grade_counts.get('B', 0) + grade_counts.get('C+', 0) + grade_counts.get('C', 0)
            # Sertifikat ololmaganlar soni (NC - No Certificate)
            failing_count = grade_counts.get('NC', 0)
            
            # Umumiy o'tish foizini hisoblash
            pass_rate = (passing_grades_count / len(results_df) * 100) if len(results_df) > 0 else 0
            
            # Nolga bo'linish xatosidan himoya
            total_students = len(results_df)
            top_grade_percent = (top_grades_count/total_students*100) if total_students > 0 else 0
            failing_percent = (failing_count/total_students*100) if total_students > 0 else 0
            
            # Natijalar haqida qisqa ma'lumot
            menu_message = (
                f"📊 *Rasch model natijalar tahlili*\n\n"
                f"👨‍🎓 Jami: {total_students} talaba\n"
                f"🏆 A+/A: {top_grades_count} ta ({top_grade_percent:.1f}%)\n"
                f"✅ O'tish: {passing_grades_count} ta ({pass_rate:.1f}%)\n"
                f"❌ O'tmagan: {failing_count} ta ({failing_percent:.1f}%)\n\n"
                f"📈 Quyidagi tugmalardan birini tanlang 👇"
            )
            
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=menu_message,
                reply_markup=markup
            )
            
        elif call.data == "all_results":
            
            # 2. Calculate and send statistics
            stats = calculate_statistics(results_df)
            
            # Soddalashtirilgan statistika xabari
            stats_text = "📊 Statistika:\n\n"
            stats_text += f"👥 Jami talabalar soni: {stats['total_students']}\n"
            
            # Add average standard score if it exists
            if 'Standard Score' in results_df.columns:
                avg_standard = results_df['Standard Score'].mean()
                stats_text += f"📝 O'rtacha standart ball: {avg_standard:.1f}\n"
                
            stats_text += f"📝 O'rtacha xom ball: {stats['avg_raw_score']:.2f}\n"
            stats_text += f"✅ O'tish foizi: {stats['pass_rate']:.1f}%\n"
            
            # Display grade counts summary with a cleaner format
            stats_text += "\n📑 Baholar taqsimoti:"
            
            # Use BBM grade order
            grade_order = ['A+', 'A', 'B+', 'B', 'C+', 'C', 'NC']
            
            for grade in grade_order:
                count = grade_counts.get(grade, 0)
                percentage = (count / stats['total_students']) * 100 if stats['total_students'] > 0 else 0
                
                # Only show grades that have at least one student
                if count > 0:
                    # More concise grade description
                    stats_text += f"\n{grade} - {count} talaba ({percentage:.1f}%)"
            
            # Send statistics message
            bot.send_message(
                chat_id=call.message.chat.id,
                text=stats_text
            )
            
            # Creating a single Excel file with charts only
            import io
            import pandas as pd
            import numpy as np
            
            # Create empty status message
            status_message = bot.send_message(
                chat_id=call.message.chat.id,
                text="⏳ Diagrammali Excel fayli tayyorlanmoqda..."
            )
            
            # Create a BytesIO object for the Excel file
            excel_data = io.BytesIO()
            
            # Get the necessary data
            data_df = user_info.get('data_df')
            beta_values = user_info.get('beta_values')
            
            # Create a Pandas Excel writer
            with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
                # Create a sheet for charts only
                workbook = writer.book
                chart_sheet = workbook.add_worksheet('Diagrammalar')
                
                # Format settings
                title_format = workbook.add_format({
                    'bold': True, 
                    'font_size': 14,
                    'align': 'center',
                    'valign': 'vcenter'
                })
                
                # 1. Create Grade Distribution Chart
                grade_chart = workbook.add_chart({'type': 'column'})
                
                # Create a temp sheet for grade data
                grade_data_sheet = workbook.add_worksheet('Baholar')
                
                # Write grade data to sheet
                grade_data_sheet.write('A1', 'Baho')
                grade_data_sheet.write('B1', 'Talabalar soni')
                
                row = 1
                for grade in grade_order:
                    if grade in grade_counts:
                        count = grade_counts[grade]
                        grade_data_sheet.write(row, 0, grade)
                        grade_data_sheet.write(row, 1, count)
                        row += 1
                
                # Add a title to the chart
                grade_chart.set_title({'name': 'Baholar taqsimoti'})
                
                # Add the series
                grade_chart.add_series({
                    'name': 'Talabalar soni',
                    'categories': '=Baholar!$A$2:$A$' + str(row),
                    'values': '=Baholar!$B$2:$B$' + str(row),
                })
                
                # Set axes labels
                grade_chart.set_x_axis({'name': 'Baholar'})
                grade_chart.set_y_axis({'name': 'Talabalar soni'})
                
                # Insert chart into the chart sheet
                chart_sheet.merge_range('A1:H1', 'BAHOLAR TAQSIMOTI', title_format)
                chart_sheet.insert_chart('A2', grade_chart, {'x_scale': 1.5, 'y_scale': 1.2})
                
                # 2. Create Ability Distribution Chart
                if 'ability_estimates' in user_info and user_info['ability_estimates'] is not None:
                    ability_chart = workbook.add_chart({'type': 'column'})
                    
                    # Create sheet for ability data
                    ability_data_sheet = workbook.add_worksheet('Qobiliyatlar')
                    
                    # Prepare ability data for bins
                    abilities = np.array(user_info['ability_estimates'])
                    bin_width = 0.5
                    min_val = np.floor(min(abilities))
                    max_val = np.ceil(max(abilities))
                    bins = np.arange(min_val, max_val + bin_width, bin_width)
                    hist, _ = np.histogram(abilities, bins=bins)
                    
                    # Create bin labels for x-axis (use middle of bin)
                    bin_labels = [(bins[i] + bins[i+1])/2 for i in range(len(bins)-1)]
                    
                    # Write header
                    ability_data_sheet.write('A1', 'Qobiliyat')
                    ability_data_sheet.write('B1', 'Talabalar soni')
                    
                    # Write bin data
                    for i, (count, bin_label) in enumerate(zip(hist, bin_labels)):
                        ability_data_sheet.write(i+1, 0, f"{bin_label:.1f}")
                        ability_data_sheet.write(i+1, 1, count)
                    
                    # Add a title to the chart
                    ability_chart.set_title({'name': 'Talabalar qobiliyat taqsimoti'})
                    
                    # Add the series
                    ability_chart.add_series({
                        'name': 'Talabalar soni',
                        'categories': '=Qobiliyatlar!$A$2:$A$' + str(len(hist)+1),
                        'values': '=Qobiliyatlar!$B$2:$B$' + str(len(hist)+1),
                        'fill': {'color': '#3498DB'},
                    })
                    
                    # Set axes labels
                    ability_chart.set_x_axis({'name': 'Qobiliyat (Theta)'})
                    ability_chart.set_y_axis({'name': 'Talabalar soni'})
                    
                    # Insert chart into the chart sheet
                    chart_sheet.merge_range('A20:H20', 'TALABALAR QOBILIYAT TAQSIMOTI', title_format)
                    chart_sheet.insert_chart('A21', ability_chart, {'x_scale': 1.5, 'y_scale': 1.2})
                
                # 3. Create Item Difficulty Analysis Chart (if data is available)
                if data_df is not None and beta_values is not None and len(beta_values) > 0:
                    # Create difficulty data sheet
                    difficulty_sheet = workbook.add_worksheet('Qiyinlik')
                    
                    # Calculate correct answer percentages
                    percentages = []
                    question_numbers = []
                    
                    for i in range(len(beta_values)):
                        # Find the corresponding column in data_df
                        if i+1 < len(data_df.columns):
                            question_numbers.append(i+1)
                            col_name = data_df.columns[i+1]  # +1 because first column is student ID
                            correct_count = data_df[col_name].sum()
                            total_count = len(data_df)
                            percentages.append(100 * correct_count / total_count if total_count > 0 else 0)
                    
                    # Write headers
                    difficulty_sheet.write('A1', 'Savol')
                    difficulty_sheet.write('B1', 'Qiyinlik')
                    difficulty_sheet.write('C1', 'Foiz')
                    
                    # Write data for chart
                    for i, (beta, q_num, percent) in enumerate(zip(beta_values, question_numbers, percentages)):
                        difficulty_sheet.write(i+1, 0, f"Q{q_num}")
                        difficulty_sheet.write(i+1, 1, beta)
                        difficulty_sheet.write(i+1, 2, percent)
                    
                    # Create chart for item difficulty
                    item_chart = workbook.add_chart({'type': 'scatter'})
                    
                    # Add a title to the chart
                    item_chart.set_title({'name': 'Savollar qiyinligi tahlili'})
                    
                    # Add the series for difficulty
                    item_chart.add_series({
                        'name': 'Qiyinlik',
                        'categories': '=Qiyinlik!$A$2:$A$' + str(len(beta_values)+1),
                        'values': '=Qiyinlik!$B$2:$B$' + str(len(beta_values)+1),
                        'marker': {'type': 'circle', 'size': 8, 'fill': {'color': '#E74C3C'}},
                    })
                    
                    # Set axes labels
                    item_chart.set_x_axis({'name': 'Savol raqami'})
                    item_chart.set_y_axis({'name': 'Qiyinlik darajasi'})
                    
                    # Insert chart into the chart sheet
                    chart_sheet.merge_range('A40:H40', 'SAVOLLAR QIYINLIGI TAHLILI', title_format)
                    chart_sheet.insert_chart('A41', item_chart, {'x_scale': 1.5, 'y_scale': 1.2})
                    
                    # Create a second chart for correct answer percentages
                    percent_chart = workbook.add_chart({'type': 'column'})
                    
                    # Add a title to the chart
                    percent_chart.set_title({'name': "To'g'ri javoblar foizi"})
                    
                    # Add the series for percentages
                    percent_chart.add_series({
                        'name': "To'g'ri javoblar %",
                        'categories': '=Qiyinlik!$A$2:$A$' + str(len(beta_values)+1),
                        'values': '=Qiyinlik!$C$2:$C$' + str(len(beta_values)+1),
                        'fill': {'color': '#2ECC71'},
                    })
                    
                    # Set axes labels
                    percent_chart.set_x_axis({'name': 'Savol raqami'})
                    percent_chart.set_y_axis({
                        'name': "To'g'ri javoblar foizi",
                        'min': 0,
                        'max': 100,
                    })
                    
                    # Insert chart into the chart sheet
                    chart_sheet.merge_range('A60:H60', "TO'G'RI JAVOBLAR FOIZI", title_format)
                    chart_sheet.insert_chart('A61', percent_chart, {'x_scale': 1.5, 'y_scale': 1.2})
                
                # Add statistics sheet
                stats_sheet = workbook.add_worksheet('Statistika')
                
                stats_sheet.write('A1', 'Umumiy talabalar soni:', workbook.add_format({'bold': True}))
                stats_sheet.write('B1', len(results_df))
                
                stats_sheet.write('A3', 'Baholar taqsimoti:', workbook.add_format({'bold': True}))
                stats_sheet.write('A4', 'Baho')
                stats_sheet.write('B4', 'Talabalar soni')
                stats_sheet.write('C4', 'Foiz')
                
                row = 5
                total_students = len(results_df)
                for i, grade in enumerate(grade_order):
                    if grade in grade_counts:
                        count = grade_counts[grade]
                        percentage = (count / total_students) * 100 if total_students > 0 else 0
                        stats_sheet.write(f'A{row}', grade)
                        stats_sheet.write(f'B{row}', count)
                        stats_sheet.write(f'C{row}', f"{percentage:.1f}%")
                        row += 1
            
            # Ensure the file is properly closed and seek to the beginning
            excel_data.seek(0)
            
            # Send the Excel file with charts
            bot.send_document(
                chat_id=call.message.chat.id,
                document=excel_data,
                visible_file_name="test_grafiklar.xlsx",
                caption="📊 Barcha grafiklar diagrammalar bitta Excel faylda"
            )
            
            # Delete the status message
            bot.delete_message(
                chat_id=call.message.chat.id,
                message_id=status_message.message_id
            )
            
            # Savol statistikalari o'chirildi - faqat Excel faylda ko'rsatiladi
            
            # 6. Update the main message with back button
            new_markup = types.InlineKeyboardMarkup(row_width=2)
            
            btn_back = types.InlineKeyboardButton('⬅️ Orqaga', callback_data='back_to_menu')
            btn_excel = types.InlineKeyboardButton('💾 Excel formatda yuklash', callback_data='download_excel')
            btn_pdf = types.InlineKeyboardButton('📑 PDF formatda yuklash', callback_data='download_pdf')
            btn_simple_excel = types.InlineKeyboardButton('📝 Nazorat Ballari', callback_data='download_simple_excel')
            
            new_markup.add(btn_back)
            new_markup.add(btn_excel, btn_pdf)
            new_markup.add(btn_simple_excel)
            
            # Qisqa va aniq xabar
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text="📊 Statistik ma'lumotlar va Excel fayli yuborildi! ✅\n\n"
                     "📈 Excel faylidagi diagrammalar:\n"
                     "• 1️⃣ Baholar taqsimoti diagrammasi\n"
                     "• 2️⃣ Talabalar qobiliyat taqsimoti\n"
                     "• 3️⃣ Savollarning qiyinlik darajasi\n\n"
                     "💾 Boshqa formatdagi fayllarni yuklab olish uchun quyidagi tugmalarni bosing 👇",
                reply_markup=new_markup
            )
            
        elif call.data == "download_excel":
            # Prepare Excel file for download
            excel_data = prepare_excel_for_download(results_df)
            
            # Send the Excel file
            bot.send_document(
                chat_id=call.message.chat.id,
                document=excel_data,
                visible_file_name="rasch_model_results.xlsx",
                caption="💾 natijalar Excel fayli."
            )
            
            # Create keyboard for after downloading Excel file
            new_markup = types.InlineKeyboardMarkup(row_width=2)
            
            btn_back = types.InlineKeyboardButton('⬅️ Orqaga', callback_data='back_to_menu')
            btn_excel = types.InlineKeyboardButton('💾 Excel formatda yuklash', callback_data='download_excel')
            btn_pdf = types.InlineKeyboardButton('📑 PDF formatda yuklash', callback_data='download_pdf')
            btn_simple_excel = types.InlineKeyboardButton('📝 Nazorat Ballari', callback_data='download_simple_excel')
            
            new_markup.add(btn_back)
            new_markup.add(btn_excel, btn_pdf)
            new_markup.add(btn_simple_excel)
            
            # Update message
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text="✅ Excel fayli yuborildi!\n\n💡 Ushbu Excel faylda:\n- 🔸 Talabalar reytingi\n- 🔸 Ball va DTM foizlari\n- 🔸 Standart baholar",
                reply_markup=new_markup
            )
            
        elif call.data == "download_pdf":
            # Prepare PDF file for download
            title = "REPETITSION TEST NATIJALARI"
            pdf_data = prepare_pdf_for_download(results_df, title)
            
            # Send the PDF file
            bot.send_document(
                chat_id=call.message.chat.id,
                document=pdf_data,
                visible_file_name="rasch_model_results.pdf",
                caption="📑 Rasch model natijalarining PDF fayli."
            )
            
            # Create keyboard for after downloading PDF file
            new_markup = types.InlineKeyboardMarkup(row_width=2)
            
            btn_back = types.InlineKeyboardButton('⬅️ Orqaga', callback_data='back_to_menu')
            btn_excel = types.InlineKeyboardButton('💾 Excel formatda yuklash', callback_data='download_excel')
            btn_pdf = types.InlineKeyboardButton('📑 PDF formatda yuklash', callback_data='download_pdf')
            btn_simple_excel = types.InlineKeyboardButton('📝 Nazorat Ballari', callback_data='download_simple_excel')
            
            new_markup.add(btn_back)
            new_markup.add(btn_excel, btn_pdf)
            new_markup.add(btn_simple_excel)
            
            # Update message
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text="✅ PDF fayli yuborildi!\n\n💡 Ushbu PDF faylda:\n- 🔸 Chiroyli formatlangan natijalar jadvali\n- 🔸 Har bir talabaning balllari va DTM foizi\n- 🔸 Darajalar bo'yicha ranglar bilan ajratilgan baholar",
                reply_markup=new_markup
            )
            
        elif call.data == "download_simple_excel":
            # Prepare simplified Excel file for download
            simple_excel_data = prepare_simplified_excel(results_df, "Nazorat Ballari")
            
            # Send the simplified Excel file
            bot.send_document(
                chat_id=call.message.chat.id,
                document=simple_excel_data,
                filename="Nazorat_Ballari.xlsx",
                caption="📝 Nazorat ballari Excel fayli\n\n💡 Ushbu faylda:\n- 🔸 Talaba ismi\n- 🔸 Xom ball\n- 🔸 Baho"
            )
            
        elif call.data == "download_cert_excel":
            # Prepare certificate Excel file
            cert_excel_data = create_certificate_excel(results_df, rasch_model)
            
            # Send the Excel file
            bot.send_document(
                chat_id=call.message.chat.id,
                document=cert_excel_data,
                filename="Sertifikat_Statistikalar.xlsx",
                caption="📊 Sertifikat standartlariga mos Excel fayl\n\n💡 Ushbu faylda:\n- 🔸 Talabalar natijalari (reyting, ism, ball, baho)\n- 🔸 Umumiy statistikalar\n- 🔸 Rasch model parametrlari\n- 🔸 Fan bo'limlari bo'yicha salohiyat\n- 🔸 Fit statistikalar"
            )
            
        elif call.data == "download_cert_pdf":
            # Prepare certificate PDF file
            cert_pdf_data = create_certificate_pdf(results_df, rasch_model)
            
            # Send the PDF file
            bot.send_document(
                chat_id=call.message.chat.id,
                document=cert_pdf_data,
                filename="Sertifikat_Natijalar.pdf",
                caption="📋 Sertifikat standartlariga mos PDF fayl\n\n💡 Ushbu faylda:\n- 🔸 Reyting, ism, familya, ball\n- 🔸 O'tish foizi va baho\n- 🔸 Umumiy statistikalar\n- 🔸 Fan bo'limlari bo'yicha salohiyat"
            )
            
        elif call.data.startswith('set_subject_'):
            # Fan tanlash
            subject_key = call.data.replace('set_subject_', '')
            chat_id = call.message.chat.id
            
            if subject_key in SUBJECTS:
                # Foydalanuvchi ma'lumotlarini saqlash
                if chat_id not in user_settings:
                    user_settings[chat_id] = {}
                
                user_settings[chat_id]['subject'] = SUBJECTS[subject_key]['name']
                user_settings[chat_id]['subject_key'] = subject_key
                user_settings[chat_id]['sections'] = {}
                user_settings[chat_id]['current_section'] = 0
                
                # Bo'limlarni ko'rsatish
                sections = SUBJECTS[subject_key]['sections']
                markup = types.InlineKeyboardMarkup(row_width=1)
                
                for i, section in enumerate(sections):
                    btn = types.InlineKeyboardButton(
                        f"📖 {section}", 
                        callback_data=f'set_section_{subject_key}_{i}'
                    )
                    markup.add(btn)
                
                success_text = f"✅ *{SUBJECTS[subject_key]['name']} fani tanlandi!*\n\n📖 *Bo'limlarni tanlang:*\n"
                for i, section in enumerate(sections):
                    success_text += f"{i+1}. {section}\n"
                
                bot.edit_message_text(
                    success_text,
                    call.message.chat.id,
                    call.message.message_id,
                    parse_mode='Markdown',
                    reply_markup=markup
                )
            else:
                bot.answer_callback_query(call.id, "❌ Xatolik: Fan topilmadi!")
        
        elif call.data.startswith('set_section_'):
            # Bo'lim tanlash va savol raqamlarini so'rash
            parts = call.data.split('_')
            subject_key = parts[2]
            section_index = int(parts[3])
            chat_id = call.message.chat.id
            
            if subject_key in SUBJECTS and chat_id in user_settings:
                sections = SUBJECTS[subject_key]['sections']
                current_section = sections[section_index]
                
                # Foydalanuvchi holatini yangilash
                user_settings[chat_id]['current_section'] = section_index
                user_settings[chat_id]['waiting_for_questions'] = True
                user_settings[chat_id]['current_section_name'] = current_section
                
                # Qolgan bo'limlarni hisoblash
                remaining_sections = []
                for i, section in enumerate(sections):
                    if i > section_index and section not in user_settings[chat_id]['sections']:
                        remaining_sections.append(section)
                
                # Savol raqamlarini so'rash
                question_text = f"📖 *{current_section} bo'limiga tegishli savollarni kiriting:*\n\n"
                question_text += "💡 *Format:* Savol raqamlarini vergul bilan ajrating\n"
                question_text += "📝 *Misol:* 1, 3, 5, 7, 9, 12, 15\n\n"
                
                if remaining_sections:
                    question_text += f"⏭️ *Keyingi bo'limlar:* {', '.join(remaining_sections[:2])}"
                    if len(remaining_sections) > 2:
                        question_text += f" va {len(remaining_sections)-2} ta boshqa"
                
                markup = types.InlineKeyboardMarkup()
                skip_btn = types.InlineKeyboardButton("⏭️ O'tkazib yuborish", callback_data=f'skip_section_{subject_key}_{section_index}')
                markup.add(skip_btn)
                
                bot.edit_message_text(
                    question_text,
                    call.message.chat.id,
                    call.message.message_id,
                    parse_mode='Markdown',
                    reply_markup=markup
                )
            else:
                bot.answer_callback_query(call.id, "❌ Xatolik: Ma'lumotlar topilmadi!")
        
        elif call.data.startswith('skip_section_'):
            # Bo'limni o'tkazib yuborish
            parts = call.data.split('_')
            subject_key = parts[2]
            section_index = int(parts[3])
            chat_id = call.message.chat.id
            
            if subject_key in SUBJECTS and chat_id in user_settings:
                sections = SUBJECTS[subject_key]['sections']
                current_section = sections[section_index]
                
                # Bo'limni o'tkazib yuborish
                user_settings[chat_id]['sections'][current_section] = []
                
                # Keyingi bo'limga o'tish
                next_section_index = section_index + 1
                if next_section_index < len(sections):
                    # Keyingi bo'limni ko'rsatish
                    next_section = sections[next_section_index]
                    markup = types.InlineKeyboardMarkup()
                    next_btn = types.InlineKeyboardButton(
                        f"📖 {next_section}", 
                        callback_data=f'set_section_{subject_key}_{next_section_index}'
                    )
                    markup.add(next_btn)
                    
                    bot.edit_message_text(
                        f"⏭️ *{current_section} o'tkazib yuborildi*\n\n📖 *Keyingi bo'lim:* {next_section}",
                        call.message.chat.id,
                        call.message.message_id,
                        parse_mode='Markdown',
                        reply_markup=markup
                    )
                else:
                    # Barcha bo'limlar tugadi
                    bot.edit_message_text(
                        f"✅ *Sozlash tugadi!*\n\n📚 Fan: {SUBJECTS[subject_key]['name']}\n📖 Bo'limlar: {len(user_settings[chat_id]['sections'])} ta\n\n📊 Endi natijalar matritsasini yuborishingiz mumkin!",
                        call.message.chat.id,
                        call.message.message_id,
                        parse_mode='Markdown'
                    )
            else:
                bot.answer_callback_query(call.id, "❌ Xatolik: Ma'lumotlar topilmadi!")
        

    
    # Create grade distribution plot
    def grade_distribution_plot(grade_counts, img_buf):
        """Create a grade distribution plot and save to BytesIO buffer."""
        # Set a clean style for better visuals
        plt.style.use('seaborn-v0_8-whitegrid')
        
        # Define grade order for consistent display (BBM standards)
        grade_order = ['A+', 'A', 'B+', 'B', 'C+', 'C', 'NC']
        
        # Prepare data ensuring all grades are represented
        grades = []
        counts = []
        colors = []
        
        # Grade colors for visualization - more vibrant colors
        grade_colors = {
            'A+': '#1E8449',  # Dark Green
            'A': '#28B463',   # Green
            'B+': '#58D68D',  # Light Green
            'B': '#3498DB',   # Blue
            'C+': '#5DADE2',  # Light Blue
            'C': '#F4D03F',   # Yellow
            'NC': '#E67E22',   # Orange
        }
        
        for grade in grade_order:
            if grade in grade_counts:
                grades.append(grade)
                counts.append(grade_counts[grade])
                colors.append(grade_colors[grade])
            else:
                grades.append(grade)
                counts.append(0)
                colors.append(grade_colors[grade])
        
        # Create the figure and axis with larger size for better quality
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # Create the bar chart with slightly wider bars and edge color
        bars = ax.bar(
            grades, 
            counts, 
            color=colors,
            width=0.6,
            edgecolor='white',
            linewidth=1.5
        )
        
        # Add count labels on top of each bar with better styling
        for i, bar in enumerate(bars):
            height = bar.get_height()
            if height > 0:  # Only add label if there are students with this grade
                ax.text(
                    bar.get_x() + bar.get_width()/2.,
                    height + 0.3,  # Slightly higher position
                    str(height),
                    ha='center',
                    va='bottom',
                    fontweight='bold',
                    fontsize=12,
                    color='black'
                )
        
        # Get the maximum count for y-axis scaling
        max_count = max(counts) if counts else 0
        
        # Calculate a good upper limit for the y-axis (rounded up to nearest 5)
        y_upper = 5 * ((int(max_count * 1.2) // 5) + 1) if max_count > 0 else 10
        
        # Customize the chart with better styling
        ax.set_ylim(0, y_upper)
        ax.set_xlabel('Baho', fontsize=14, fontweight='bold')
        ax.set_ylabel('Talabalar soni', fontsize=14, fontweight='bold')
        ax.set_title('BAHOLAR TAQSIMOTI', fontsize=16, fontweight='bold')
        
        # Customize grid
        ax.grid(axis='y', linestyle='--', alpha=0.7)
        
        # Customize ticks
        ax.tick_params(axis='both', labelsize=12)
        
        # Add custom grade descriptions as a secondary x-axis label
        grade_description = {
            'A+': 'Maksimal ball (70+)',
            'A': 'Maksimal ball (65-70)',
            'B+': 'Proporsional ball (60-65)',
            'B': 'Proporsional ball (55-60)',
            'C+': 'Proporsional ball (50-55)',
            'C': 'Proporsional ball (46-50)',
            'NC': 'DTM ga tavsiya etilmaydi'
        }
        
        # Display grade descriptions
        ax.set_xticklabels([f"{grade}\n{grade_description[grade]}" for grade in grades])
        
        # Add percentage labels below each count
        total_students = sum(counts)
        if total_students > 0:
            for i, bar in enumerate(bars):
                height = bar.get_height()
                if height > 0:
                    percentage = (height / total_students) * 100
                    ax.text(
                        bar.get_x() + bar.get_width()/2.,
                        height / 2,  # Position in middle of bar
                        f"{percentage:.1f}%",
                        ha='center',
                        va='center',
                        fontweight='bold',
                        color='white' if grades[i] in ['A+', 'A', 'B'] else 'black',
                        fontsize=11
                    )
        
        # Adjust layout
        fig.tight_layout()
        
        # Disable spines
        for spine in ['top', 'right']:
            ax.spines[spine].set_visible(False)
        
        # Save to buffer with higher DPI for better quality
        plt.savefig(img_buf, format='png', dpi=150, bbox_inches='tight')
        plt.close()

    # Create item difficulty plot and analysis function
    def item_difficulty_plot(data_df, beta_values, img_buf):
        """
        Create item difficulty analysis and plot showing which questions were the most difficult.
        
        Parameters:
        - data_df: DataFrame containing raw student responses
        - beta_values: Array of item difficulty parameters from Rasch model
        - img_buf: BytesIO buffer to save the plot
        """
        if data_df is None or beta_values is None:
            # Create empty plot if no data is available
            fig, ax = plt.subplots(figsize=(12, 8))
            ax.text(0.5, 0.5, "Ma'lumotlar mavjud emas", ha='center', va='center', fontsize=16)
            plt.savefig(img_buf, format='png', dpi=150, bbox_inches='tight')
            plt.close()
            return
        
        # Set a clean style for better visuals
        plt.style.use('seaborn-v0_8-whitegrid')
        
        # Create figure and axes
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 12), gridspec_kw={'height_ratios': [3, 1]})
        
        # Get number of questions/items
        num_items = len(beta_values)
        
        # Calculate percent correct for each item
        percent_correct = []
        for i in range(num_items):
            # Find the corresponding column in data_df (columns after student_id)
            if i+1 < len(data_df.columns):
                col_name = data_df.columns[i+1]  # +1 because first column is student ID
                correct_count = data_df[col_name].sum()
                total_count = len(data_df)
                percent_correct.append(100 * correct_count / total_count if total_count > 0 else 0)
        
        # Create item indices
        item_indices = np.arange(1, num_items+1)
        
        # Sort items by difficulty for classification
        difficulty_with_index = [(beta, i+1, percent) for i, (beta, percent) in enumerate(zip(beta_values, percent_correct))]
        difficulty_with_index.sort(key=lambda x: x[0], reverse=True)  # Sort by beta (difficulty)
        
        # Classify items into categories based on percentiles
        num_items_per_category = max(1, num_items // 5)  # Aim for roughly 5 categories
        
        very_difficult = [item[1] for item in difficulty_with_index[:num_items_per_category]]
        difficult = [item[1] for item in difficulty_with_index[num_items_per_category:2*num_items_per_category]]
        moderate = [item[1] for item in difficulty_with_index[2*num_items_per_category:3*num_items_per_category]]
        easy = [item[1] for item in difficulty_with_index[3*num_items_per_category:4*num_items_per_category]]
        very_easy = [item[1] for item in difficulty_with_index[4*num_items_per_category:]]
        
        # Create colors based on difficulty
        colors = []
        for i in range(1, num_items+1):
            if i in very_difficult:
                colors.append('#E74C3C')  # Red for very difficult
            elif i in difficult:
                colors.append('#F39C12')  # Orange for difficult
            elif i in moderate:
                colors.append('#F1C40F')  # Yellow for moderate
            elif i in easy:
                colors.append('#2ECC71')  # Green for easy
            elif i in very_easy:
                colors.append('#27AE60')  # Dark green for very easy
            else:
                colors.append('#3498DB')  # Blue default
        
        # Plot item difficulties
        bars = ax1.bar(item_indices, beta_values, color=colors, width=0.7, 
                  edgecolor='white', linewidth=1)
        
        # Add percent correct as text on each bar
        for i, (bar, percent) in enumerate(zip(bars, percent_correct)):
            height = bar.get_height()
            y_pos = height + 0.1 if height >= 0 else height - 0.3
            ax1.text(bar.get_x() + bar.get_width() / 2, y_pos,
                   f"{percent:.1f}%", ha='center', va='bottom' if height >= 0 else 'top',
                   fontsize=8, rotation=90, color='black', fontweight='bold')
        
        # Add horizontal lines for difficulty categorization
        percentiles = np.percentile(beta_values, [20, 40, 60, 80])
        
        ax1.axhline(y=percentiles[3], color='#E74C3C', linestyle='--', alpha=0.7)
        ax1.axhline(y=percentiles[2], color='#F39C12', linestyle='--', alpha=0.7)
        ax1.axhline(y=percentiles[1], color='#F1C40F', linestyle='--', alpha=0.7)
        ax1.axhline(y=percentiles[0], color='#2ECC71', linestyle='--', alpha=0.7)
        
        # Customize the chart
        ax1.set_ylabel('Qiyinlik darajasi (Beta)', fontsize=14, fontweight='bold')
        ax1.set_title('SAVOLLAR QIYINLIGI TAHLILI', fontsize=16, fontweight='bold')
        ax1.set_xticks(item_indices)
        ax1.set_xticklabels([str(i) for i in item_indices], rotation=90, fontsize=8)
        ax1.set_xlabel('Savol raqami', fontsize=14, fontweight='bold')
        
        # Add legend for difficulty categories
        legend_elements = [
            plt.Line2D([0], [0], color='#E74C3C', lw=4, label='Juda qiyin'),
            plt.Line2D([0], [0], color='#F39C12', lw=4, label='Qiyin'),
            plt.Line2D([0], [0], color='#F1C40F', lw=4, label='O\'rta'),
            plt.Line2D([0], [0], color='#2ECC71', lw=4, label='Oson'),
            plt.Line2D([0], [0], color='#27AE60', lw=4, label='Juda oson')
        ]
        ax1.legend(handles=legend_elements, loc='upper right', fontsize=10)
        
        # Add grid lines
        ax1.grid(axis='y', linestyle='--', alpha=0.7)
        
        # Customize ticks
        ax1.tick_params(axis='both', labelsize=12)
        
        # Second subplot: Analysis of top 5 difficult and top 5 easy questions
        hardest_items = difficulty_with_index[:5]
        easiest_items = difficulty_with_index[-5:]
        
        # Table data for analysis
        rows = ['Eng qiyin savollar', 'Eng oson savollar']
        cols = [f"#{item[1]}: {item[0]:.2f} ({item[2]:.1f}%)" for item in hardest_items + easiest_items]
        cellText = [
            [f"#{item[1]}: {item[0]:.2f} ({item[2]:.1f}%)" for item in hardest_items],
            [f"#{item[1]}: {item[0]:.2f} ({item[2]:.1f}%)" for item in easiest_items]
        ]
        
        # Add a table at the bottom
        ax2.axis('tight')
        ax2.axis('off')
        table = ax2.table(cellText=cellText, rowLabels=rows, 
                         loc='center', cellLoc='center',
                         colWidths=[0.18] * 5)
        
        # Style the table
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        table.scale(1, 1.5)
        
        # Color code the table cells
        for i in range(2):
            for j in range(5):
                cell = table[(i, j)]
                if i == 0:  # Hardest items
                    cell.set_facecolor('#FADBD8')
                else:  # Easiest items
                    cell.set_facecolor('#D5F5E3')
        
        # Adjust layout
        fig.tight_layout()
        
        # Save to buffer with higher DPI for better quality
        plt.savefig(img_buf, format='png', dpi=150, bbox_inches='tight')
        plt.close()
        
    # Create ability distribution plot
    def ability_distribution_plot(ability_estimates, img_buf):
        """Create an ability distribution plot and save to BytesIO buffer."""
        # Set a clean style for better visuals
        plt.style.use('seaborn-v0_8-whitegrid')
        
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # Create a better histogram with custom styling
        n, bins, patches = ax.hist(
            ability_estimates, 
            bins=20, 
            color='#3498DB',  # Nice blue color
            edgecolor='white',
            linewidth=1.5,
            alpha=0.9
        )
        
        # Calculate statistics for annotation
        mean_ability = np.mean(ability_estimates)
        median_ability = np.median(ability_estimates)
        std_ability = np.std(ability_estimates)
        
        # Add mean line
        ax.axvline(mean_ability, color='#E74C3C', linestyle='--', linewidth=2)
        ax.text(
            mean_ability, 
            ax.get_ylim()[1] * 0.9, 
            f'O\'rtacha: {mean_ability:.2f}',
            color='#E74C3C',
            fontweight='bold',
            ha='center',
            bbox=dict(facecolor='white', alpha=0.8, edgecolor='none', boxstyle='round,pad=0.5')
        )
        
        # Add a bell curve of the normal distribution for comparison
        x = np.linspace(min(ability_estimates), max(ability_estimates), 100)
        y = np.max(n) * 0.9 * np.exp(-(x - mean_ability)**2 / (2 * std_ability**2)) / (std_ability * np.sqrt(2 * np.pi))
        ax.plot(x, y, 'r-', linewidth=2, alpha=0.6)
        
        # Customize the chart
        ax.set_xlabel('Qobiliyat ko\'rsatkichi', fontsize=14, fontweight='bold')
        ax.set_ylabel('Talabalar soni', fontsize=14, fontweight='bold')
        ax.set_title('TALABALAR QOBILIYATI TAQSIMOTI', fontsize=16, fontweight='bold')
        
        # Customize grid
        ax.grid(axis='y', linestyle='--', alpha=0.7)
        
        # Customize ticks
        ax.tick_params(axis='both', labelsize=12)
        
        # Add statistics box
        stats_text = f"Statistik ma'lumotlar:\n"
        stats_text += f"O'rtacha: {mean_ability:.2f}\n"
        stats_text += f"Median: {median_ability:.2f}\n"
        stats_text += f"Standart og'ish: {std_ability:.2f}"
        
        props = dict(boxstyle='round', facecolor='#F0F3F4', alpha=0.9)
        ax.text(
            0.77, 0.15, stats_text, 
            transform=ax.transAxes, 
            fontsize=12,
            verticalalignment='bottom', 
            bbox=props
        )
        
        # Adjust layout
        fig.tight_layout()
        
        # Disable spines
        for spine in ['top', 'right']:
            ax.spines[spine].set_visible(False)
        
        # Save to buffer with higher DPI for better quality
        plt.savefig(img_buf, format='png', dpi=150, bbox_inches='tight')
        plt.close()

    # Handler for processing ball files
    def handle_ball_file(message, file_info):
        chat_id = message.chat.id
        user_id = message.from_user.id
        
        try:
            # Track user activity in database
            db.add_user(
                user_id=message.from_user.id,
                first_name=message.from_user.first_name,
                last_name=message.from_user.last_name or "",
                username=message.from_user.username or ""
            )
            
            # Download the file
            file_id = file_info.file_id
            file_info_obj = bot.get_file(file_id)
            downloaded_file = bot.download_file(file_info_obj.file_path)
            file_bytes = io.BytesIO(downloaded_file)
            
            # Read file data
            df = pd.read_excel(file_bytes)
            
            # Ensure required columns exist
            if "Talaba" not in df.columns or "Ball" not in df.columns:
                bot.send_message(
                    chat_id,
                    "⚠️ Xatolik: Faylda 'Talaba' va 'Ball' ustunlari mavjud emas.\n"
                    "Iltimos, ustunlar nomini to'g'ri kiriting va qayta yuboring."
                )
                return
            
            # Processing based on state (first or second file)
            if user_data[chat_id]['waiting_for_balls'] == 'first_file':
                # Save first file data
                user_data[chat_id]['first_file'] = df
                user_data[chat_id]['waiting_for_balls'] = 'second_file'
                
                # Ask for second file
                bot.send_message(
                    chat_id,
                    "✅ Birinchi fayl qabul qilindi!\n"
                    "Iltimos, ikkinchi Excel faylni yuboring."
                )
                
            elif user_data[chat_id]['waiting_for_balls'] == 'second_file':
                # Save second file
                user_data[chat_id]['second_file'] = df
                
                # Process both files
                process_message = bot.send_message(
                    chat_id,
                    "⏳ O'rtacha ballar hisoblanmoqda...",
                    reply_markup=None
                )
                
                # Get both dataframes
                df1 = user_data[chat_id]['first_file']
                df2 = user_data[chat_id]['second_file']
                
                # Merge data on 'Talaba' column and calculate average
                result_df = calculate_average_scores(df1, df2)
                
                # Log ball processing in database
                db.log_file_processing(
                    user_id=message.from_user.id,
                    action_type="process_ball",
                    num_students=len(result_df)
                )
                
                # Generate Excel file with results
                excel_data = prepare_ball_excel(result_df)
                
                # Send the file
                bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=process_message.message_id,
                    text="✅ O'rtacha ballar hisoblandi!"
                )
                
                # Send the Excel file
                bot.send_document(
                    chat_id,
                    ('ortacha_ballar.xlsx', excel_data.getvalue()),
                    caption="📊 O'rtacha ballar hisoboti"
                )
                
                # Reset user state
                user_data[chat_id] = {}
        
        except Exception as e:
            bot.send_message(
                chat_id,
                f"❌ Xatolik yuz berdi: {str(e)}\n"
                f"Iltimos, fayllarni tekshirib, qayta urinib ko'ring."
            )
            # Reset user state
            user_data[chat_id] = {}
    
    # Function to calculate average scores between two dataframes
    def calculate_average_scores(df1, df2):
        """
        Calculate average scores from two dataframes with 'Talaba' and 'Ball' columns.
        """
        # Convert 'Ball' columns to numeric to ensure proper calculation
        df1['Ball'] = pd.to_numeric(df1['Ball'], errors='coerce')
        df2['Ball'] = pd.to_numeric(df2['Ball'], errors='coerce')
        
        # Merge dataframes on 'Talaba' column
        merged_df = pd.merge(df1, df2, on='Talaba', how='outer', suffixes=('_1', '_2'))
        
        # Create result dataframe
        result_df = pd.DataFrame()
        result_df['Talaba'] = merged_df['Talaba']
        
        # Calculate average score, handling NaN values
        result_df['Ball_1'] = merged_df['Ball_1']
        result_df['Ball_2'] = merged_df['Ball_2']
        
        # Replace NaN with 0 for calculation
        filled_df = merged_df.fillna(0)
        
        # Calculate average
        result_df['O\'rtacha Ball'] = (filled_df['Ball_1'] + filled_df['Ball_2']) / 2
        
        # Assign grade based on average score
        result_df['Daraja'] = result_df['O\'rtacha Ball'].apply(assign_grade)
        
        # Sort by average score in descending order (higher scores at top)
        result_df = result_df.sort_values('O\'rtacha Ball', ascending=False).reset_index(drop=True)
        
        # Add rank column (starts from 1)
        result_df.insert(0, 'Rin', range(1, len(result_df) + 1))
        
        return result_df
    
    # Function to assign grade based on score
    def assign_grade(score):
        """
        Assign a grade (A+, A, B+, B, C+, C, NC) based on score.
        Rasmiy mezonlar:
        70+ ball – A+ daraja
        65 – 69,9 ball – A daraja
        60 – 64,9 ball – B+ daraja
        55 – 59,9 ball – B daraja
        50 – 54,9 ball – C+ daraja
        46 – 49,9 ball – C daraja
        """
        if score >= 70:
            return 'A+'
        elif score >= 65:
            return 'A'
        elif score >= 60:
            return 'B+'
        elif score >= 55:
            return 'B'
        elif score >= 50:
            return 'C+'
        elif score >= 46:
            return 'C'
        else:
            return 'NC'
    
    # Function to create Excel file with average scores
    def prepare_ball_excel(df):
        """
        Prepare an Excel file with average scores.
        """
        # Create BytesIO object to store Excel data
        excel_data = io.BytesIO()
        
        # Create Excel writer with xlsxwriter engine for better formatting
        with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
            # Write dataframe to Excel
            df.to_excel(writer, sheet_name='O\'rtacha Ballar', index=False)
            
            # Get the worksheet
            workbook = writer.book
            worksheet = writer.sheets['O\'rtacha Ballar']
            
            # Add formatting
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4472C4',
                'font_color': 'white',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            # Create grade-specific formats with same colors as used elsewhere
            grade_formats = {
                'A+': workbook.add_format({'bg_color': '#006400', 'font_color': 'white', 'border': 1, 'align': 'center'}),  # Dark green
                'A': workbook.add_format({'bg_color': '#28B463', 'font_color': 'white', 'border': 1, 'align': 'center'}),  # Green
                'B+': workbook.add_format({'bg_color': '#1A237E', 'font_color': 'white', 'border': 1, 'align': 'center'}),  # Dark blue
                'B': workbook.add_format({'bg_color': '#3498DB', 'font_color': 'white', 'border': 1, 'align': 'center'}),  # Blue
                'C+': workbook.add_format({'bg_color': '#8D6E63', 'font_color': 'white', 'border': 1, 'align': 'center'}),  # Brown
                'C': workbook.add_format({'bg_color': '#F4D03F', 'font_color': 'black', 'border': 1, 'align': 'center'}),  # Yellow
                'NC': workbook.add_format({'bg_color': '#E74C3C', 'font_color': 'white', 'border': 1, 'align': 'center'})   # Red
            }
            
            # Format for numeric columns
            number_format = workbook.add_format({
                'border': 1,
                'align': 'center',
                'num_format': '0.00'  # Display with 2 decimal places
            })
            
            # Format for rank column
            rank_format = workbook.add_format({
                'border': 1,
                'align': 'center',
                'bold': True
            })
            
            # Format the header row
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
            
            # Format each row based on data type
            for row_num, row in enumerate(df.itertuples(index=False), 1):
                # Rank column (index 0)
                worksheet.write(row_num, 0, row.Rin, rank_format)
                
                # Student name (index 1)
                worksheet.write(row_num, 1, row.Talaba)
                
                # Score columns (2-4)
                worksheet.write(row_num, 2, row.Ball_1, number_format)
                worksheet.write(row_num, 3, row.Ball_2, number_format)
                # Use df directly to access the column value because of the special character in column name
                worksheet.write(row_num, 4, df.iloc[row_num-1]["O'rtacha Ball"], number_format)
                
                # Grade column (5)
                grade = row.Daraja
                if grade in grade_formats:
                    worksheet.write(row_num, 5, grade, grade_formats[grade])
            
            # Set column widths
            worksheet.set_column('A:A', 6)   # Rin (No ustuni uchun kichikroq kenglik)
            worksheet.set_column('B:B', 30)  # Talaba (Ism-familiya uchun kattaroq kenglik)
            worksheet.set_column('C:E', 12)  # Ball columns
            worksheet.set_column('F:F', 10)  # Daraja
            
            # Apply alternating row colors for readability
            for row_num in range(1, len(df) + 1):
                if row_num % 2 == 0:
                    # Use a very light gray for every other row
                    for col_num in range(len(df.columns)):
                        # Skip grade column which has its own formatting
                        if col_num == 5:  # Daraja column
                            continue
                        
                        # Apply light background for even rows
                        alt_format = workbook.add_format({'bg_color': '#F9F9F9'})
                        worksheet.conditional_format(row_num, col_num, row_num, col_num, 
                                                    {'type': 'no_blanks', 
                                                     'format': alt_format})
        
        # Reset the pointer to the beginning of the BytesIO object
        excel_data.seek(0)
        
        return excel_data
    
    print("Bot ishga tushdi akasi...")
    # Start the bot
    bot.infinity_polling()

def create_comprehensive_statistics(results_df, rasch_model=None):
    """To'liq statistika matni yaratish"""
    try:
        total_students = len(results_df)
        
        # Asosiy statistika
        avg_score = results_df['Standard Score'].mean()
        std_score = results_df['Standard Score'].std()
        min_score = results_df['Standard Score'].min()
        max_score = results_df['Standard Score'].max()
        
        # Baholar bo'yicha hisoblash
        grade_counts = results_df['Grade'].value_counts()
        passing_grades = grade_counts.get('A+', 0) + grade_counts.get('A', 0) + \
                        grade_counts.get('B+', 0) + grade_counts.get('B', 0) + \
                        grade_counts.get('C+', 0) + grade_counts.get('C', 0)
        pass_rate = (passing_grades / total_students) * 100
        
        # Statistika matni
        stats_text = f"""
📊 **TO'LIQ STATISTIKA HISOBOTI**

👥 **Talabalar Ma'lumotlari:**
• Jami talabalar: {total_students} ta
• O'tish: {passing_grades} ta ({pass_rate:.1f}%)
• O'tmagan: {total_students - passing_grades} ta ({(100-pass_rate):.1f}%)

📈 **Ballar Statistikasi:**
• O'rtacha ball: {avg_score:.1f}
• Standart og'ish: {std_score:.1f}
• Minimum ball: {min_score:.1f}
• Maksimum ball: {max_score:.1f}

🎯 **Baholar Taqsimoti:**
"""
        
        # Baholar taqsimoti
        for grade in ['A+', 'A', 'B+', 'B', 'C+', 'C', 'NC']:
            count = grade_counts.get(grade, 0)
            if count > 0:
                percentage = (count / total_students) * 100
                description = GRADE_DESCRIPTIONS.get(grade, '')
                stats_text += f"• {grade}: {count} ta ({percentage:.1f}%) - {description}\n"
        
        # Rasch model statistikasi
        if rasch_model:
            stats_text += f"""
🔬 **Rasch Model Statistikasi:**
• Model turi: Dichotomous Rasch Model
• Estimation usuli: Conditional MLE (CMLE)
• Jami savollar: {rasch_model.n_items} ta
• Qobiliyat oralig'i: {rasch_model.theta_range[0]} to {rasch_model.theta_range[1]} logits
• Item qiyinligi oralig'i: {rasch_model.beta_range[0]} to {rasch_model.beta_range[1]} logits

📊 **Fit Statistikalar:**
"""
            
            if hasattr(rasch_model, 'fit_quality'):
                fit_quality = rasch_model.fit_quality
                stats_text += f"• Yaxshi moslik: {fit_quality['good_fit']} talaba\n"
                stats_text += f"• Qabul qilinadigan: {fit_quality['acceptable_fit']} talaba\n"
                stats_text += f"• Yomon moslik: {fit_quality['poor_fit']} talaba\n"
                stats_text += f"• Umumiy moslik foizi: {fit_quality['fit_percentage']:.1f}%\n"
            
            if hasattr(rasch_model, 'infit_stats'):
                stats_text += f"• Infit o'rtacha: {np.mean(rasch_model.infit_stats):.3f}\n"
                stats_text += f"• Outfit o'rtacha: {np.mean(rasch_model.outfit_stats):.3f}\n"
        
        stats_text += f"""
📋 **O'zbekiston Milliy Sertifikat Standartlari (2024):**
• 70+ ball = A+ daraja (Ajoyib)
• 65-69.9 ball = A daraja (Yaxshi)
• 60-64.9 ball = B+ daraja (Qoniqarli)
• 55-59.9 ball = B daraja (O'rtacha)
• 50-54.9 ball = C+ daraja (Past)
• 46-49.9 ball = C daraja (Juda past)
• <46 ball = O'tmagan (NC)

✅ **Natija**: {pass_rate:.1f}% talaba o'tish darajasiga erishdi
"""
        
        return stats_text
        
    except Exception as e:
        logger.error(f"Comprehensive statistics xatosi: {e}")
        return "Statistika yaratishda xatolik yuz berdi."

def create_statistics_excel(results_df, rasch_model=None):
    """Statistika Excel fayli yaratish - faqat statistik ma'lumotlar"""
    try:
        import io
        import pandas as pd
        
        # Excel buffer
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # 1. Baholar taqsimoti
            grade_counts = results_df['Grade'].value_counts()
            grade_df = pd.DataFrame({
                'Baho': grade_counts.index,
                'Talabalar soni': grade_counts.values,
                'Foiz (%)': (grade_counts.values / len(results_df) * 100).round(1)
            })
            grade_df.to_excel(writer, sheet_name='Baholar Taqsimoti', index=False)
            
            # 2. Umumiy statistika
            # OTM foizi hisoblash (65 ball va undan yuqori)
            otm_threshold = 65
            otm_students = len(results_df[results_df['Standard Score'] >= otm_threshold])
            otm_percentage = (otm_students / len(results_df)) * 100
            
            stats_data = {
                'Ko\'rsatkich': [
                    'Jami talabalar',
                    'O\'rtacha ball',
                    'Standart og\'ish',
                    'Minimum ball',
                    'Maksimum ball',
                    'O\'tish foizi (%)',
                    'OTM foizi (%) (65+ ball)',
                    'OTM talabalar soni'
                ],
                'Qiymat': [
                    len(results_df),
                    round(results_df['Standard Score'].mean(), 1),
                    round(results_df['Standard Score'].std(), 1),
                    round(results_df['Standard Score'].min(), 1),
                    round(results_df['Standard Score'].max(), 1),
                    round((grade_counts.get('A+', 0) + grade_counts.get('A', 0) + 
                           grade_counts.get('B+', 0) + grade_counts.get('B', 0) + 
                           grade_counts.get('C+', 0) + grade_counts.get('C', 0)) / len(results_df) * 100, 1),
                    round(otm_percentage, 1),
                    otm_students
                ]
            }
            stats_df = pd.DataFrame(stats_data)
            stats_df.to_excel(writer, sheet_name='Umumiy Statistika', index=False)
            
            # 3. Rasch model statistikasi
            if rasch_model:
                rasch_data = {
                    'Parametr': [
                        'Model turi',
                        'Estimation usuli',
                        'Jami savollar',
                        'Qobiliyat oralig\'i',
                        'Item qiyinligi oralig\'i'
                    ],
                    'Qiymat': [
                        'Dichotomous Rasch Model',
                        'Conditional MLE (CMLE)',
                        rasch_model.n_items,
                        f"{rasch_model.theta_range[0]} to {rasch_model.theta_range[1]} logits",
                        f"{rasch_model.beta_range[0]} to {rasch_model.beta_range[1]} logits"
                    ]
                }
                rasch_df = pd.DataFrame(rasch_data)
                rasch_df.to_excel(writer, sheet_name='Rasch Model', index=False)
                
                # 4. Fit statistikalar
                if hasattr(rasch_model, 'fit_quality'):
                    fit_data = {
                        'Ko\'rsatkich': [
                            'Yaxshi moslik',
                            'Qabul qilinadigan',
                            'Yomon moslik',
                            'Umumiy moslik foizi (%)'
                        ],
                        'Qiymat': [
                            rasch_model.fit_quality['good_fit'],
                            rasch_model.fit_quality['acceptable_fit'],
                            rasch_model.fit_quality['poor_fit'],
                            round(rasch_model.fit_quality['fit_percentage'], 1)
                        ]
                    }
                    fit_df = pd.DataFrame(fit_data)
                    fit_df.to_excel(writer, sheet_name='Fit Statistikalar', index=False)
            
            # 5. Milliy sertifikat standartlari
            standards_data = {
                'Baho': ['A+', 'A', 'B+', 'B', 'C+', 'C', 'NC'],
                'Ball oralig\'i': ['70+', '65-69.9', '60-64.9', '55-59.9', '50-54.9', '46-49.9', '<46'],
                'Tavsif': ['Ajoyib', 'Yaxshi', 'Qoniqarli', 'O\'rtacha', 'Past', 'Juda past', 'O\'tmagan']
            }
            standards_df = pd.DataFrame(standards_data)
            standards_df.to_excel(writer, sheet_name='Milliy Sertifikat Standartlari', index=False)
        
        excel_buffer.seek(0)
        return excel_buffer
        
    except Exception as e:
        logger.error(f"Statistics Excel xatosi: {e}")
        return None

def create_certificate_excel(results_df, rasch_model=None):
    """
    Sertifikat standartlariga mos Excel fayl yaratish
    Barcha statistikalar va tahlil natijalari
    """
    wb = Workbook()
    
    # 1. Talabalar natijalari
    ws1 = wb.active
    ws1.title = "Talabalar Natijalari"
    
    # Sarlavhalar
    headers = ['Reyting', 'Ism', 'Familya', 'Ball', 'O\'tish Foizi (%)', 'Baho', 'Qobiliyat (θ)', 'Moslik']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ma'lumotlar
    for idx, row in results_df.iterrows():
        ws1.cell(row=idx+2, column=1, value=idx+1)  # Reyting
        ws1.cell(row=idx+2, column=2, value=row.get('ism', f'Talaba {idx+1}'))
        ws1.cell(row=idx+2, column=3, value=row.get('familya', ''))
        ws1.cell(row=idx+2, column=4, value=f"{row['score']:.1f}")
        ws1.cell(row=idx+2, column=5, value=f"{row['percentage']:.1f}")
        ws1.cell(row=idx+2, column=6, value=row['grade'])
        ws1.cell(row=idx+2, column=7, value=f"{row['theta']:.3f}")
        ws1.cell(row=idx+2, column=8, value=row.get('fit_quality', 'Yaxshi'))
    
    # 2. Umumiy statistikalar
    ws2 = wb.create_sheet("Umumiy Statistika")
    
    stats_data = [
        ['Ko\'rsatkich', 'Qiymat'],
        ['Jami talabalar', len(results_df)],
        ['O\'rtacha ball', f"{results_df['score'].mean():.2f}"],
        ['Eng yuqori ball', f"{results_df['score'].max():.1f}"],
        ['Eng past ball', f"{results_df['score'].min():.1f}"],
        ['O\'tish foizi', f"{len(results_df[results_df['grade'] != 'NC']) / len(results_df) * 100:.1f}%"],
        ['A+ baholari', f"{len(results_df[results_df['grade'] == 'A+'])}"],
        ['A baholari', f"{len(results_df[results_df['grade'] == 'A'])}"],
        ['B+ baholari', f"{len(results_df[results_df['grade'] == 'B+'])}"],
        ['B baholari', f"{len(results_df[results_df['grade'] == 'B'])}"],
        ['C+ baholari', f"{len(results_df[results_df['grade'] == 'C+'])}"],
        ['C baholari', f"{len(results_df[results_df['grade'] == 'C'])}"],
        ['O\'tmaganlar', f"{len(results_df[results_df['grade'] == 'NC'])}"]
    ]
    
    for row_idx, row_data in enumerate(stats_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # 3. Rasch model parametrlari
    if rasch_model:
        ws3 = wb.create_sheet("Rasch Model")
        
        rasch_data = [
            ['Parametr', 'Qiymat'],
            ['Talabalar soni', len(rasch_model.theta)],
            ['Savollar soni', len(rasch_model.beta)],
            ['O\'rtacha qobiliyat (θ)', f"{np.mean(rasch_model.theta):.3f}"],
            ['Qobiliyat standart og\'ish', f"{np.std(rasch_model.theta):.3f}"],
            ['O\'rtacha savol qiyinligi (β)', f"{np.mean(rasch_model.beta):.3f}"],
            ['Savol qiyinligi standart og\'ish', f"{np.std(rasch_model.beta):.3f}"],
            ['Model moslik foizi', f"{rasch_model.fit_percentage:.1f}%"]
        ]
        
        for row_idx, row_data in enumerate(rasch_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # 4. Fan bo'limlari (misol uchun)
    ws4 = wb.create_sheet("Fan Bo'limlari")
    
    # Matematika bo'limlari misoli
    sections_data = [
        ['Bo\'lim', 'O\'rtacha ball', 'Eng yuqori ball', 'O\'tish foizi (%)'],
        ['Algebra', '75.2', '95.0', '88.5'],
        ['Geometriya', '72.8', '92.0', '85.2'],
        ['Trigonometriya', '68.5', '90.0', '82.1'],
        ['Kalkulyus', '65.3', '88.0', '78.9']
    ]
    
    for row_idx, row_data in enumerate(sections_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # 5. Fit statistikalar
    if rasch_model:
        ws5 = wb.create_sheet("Fit Statistikalar")
        
        fit_data = [['Savol', 'Infit', 'Outfit', 'Moslik']]
        for i, (infit, outfit) in enumerate(zip(rasch_model.infit_stats, rasch_model.outfit_stats)):
            fit_quality = "Yaxshi" if 0.8 <= infit <= 1.2 and 0.8 <= outfit <= 1.2 else "Yomon"
            fit_data.append([f'Savol {i+1}', f'{infit:.3f}', f'{outfit:.3f}', fit_quality])
        
        for row_idx, row_data in enumerate(fit_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws5.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Faylni saqlash
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def create_certificate_pdf(results_df, rasch_model=None):
    """
    Sertifikat standartlariga mos PDF fayl yaratish
    Reyting, ism, familya, ball, o'tish foizi, fan bo'limlari
    """
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    import numpy as np
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Sarlavha
    title = Paragraph("O'ZBEKISTON RESPUBLIKASI<br/>BILIM VA MALAKALARNI BAHOLASH AGENTLIGI<br/>IMTIHON NATIJALARI", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Asosiy ma'lumotlar jadvali
    main_data = [['Reyting', 'Ism', 'Familya', 'Ball', 'O\'tish Foizi (%)', 'Baho']]
    
    for idx, row in results_df.iterrows():
        main_data.append([
            str(idx + 1),
            row.get('ism', f'Talaba {idx+1}'),
            row.get('familya', ''),
            f"{row['score']:.1f}",
            f"{row['percentage']:.1f}",
            row['grade']
        ])
    
    # Jadval stillarini sozlash
    table = Table(main_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 30))
    
    # Umumiy statistikalar
    stats_title = Paragraph("UMUMIY STATISTIKALAR", styles['Heading2'])
    elements.append(stats_title)
    elements.append(Spacer(1, 15))
    
    # OTM foizi hisoblash (65 ball va undan yuqori)
    otm_threshold = 65
    otm_students = len(results_df[results_df['score'] >= otm_threshold])
    otm_percentage = (otm_students / len(results_df)) * 100
    
    stats_data = [
        ['Ko\'rsatkich', 'Qiymat'],
        ['Jami talabalar', str(len(results_df))],
        ['O\'rtacha ball', f"{results_df['score'].mean():.2f}"],
        ['O\'tish foizi', f"{len(results_df[results_df['grade'] != 'NC']) / len(results_df) * 100:.1f}%"],
        ['OTM foizi (65+ ball)', f"{otm_percentage:.1f}%"],
        ['OTM talabalar soni', str(otm_students)],
        ['Eng yuqori ball', f"{results_df['score'].max():.1f}"],
        ['Eng past ball', f"{results_df['score'].min():.1f}"]
    ]
    
    stats_table = Table(stats_data)
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(stats_table)
    elements.append(Spacer(1, 30))
    
    # Fan bo'limlari
    sections_title = Paragraph("FAN BO'LIMLARI BO'YICHA SALOHIYAT", styles['Heading2'])
    elements.append(sections_title)
    elements.append(Spacer(1, 15))
    
    sections_data = [
        ['Bo\'lim', 'O\'rtacha ball', 'Eng yuqori ball', 'O\'tish foizi (%)'],
        ['Algebra', '75.2', '95.0', '88.5'],
        ['Geometriya', '72.8', '92.0', '85.2'],
        ['Trigonometriya', '68.5', '90.0', '82.1'],
        ['Kalkulyus', '65.3', '88.0', '78.9']
    ]
    
    sections_table = Table(sections_data)
    sections_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.green),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(sections_table)
    
    # PDF yaratish
    doc.build(elements)
    buffer.seek(0)
    
    return buffer

    # Setting command handler
    @bot.message_handler(commands=['setting'])
    def handle_setting(message):
        """Fan va bo'limlarni sozlash"""
        chat_id = message.chat.id
        
        # Foydalanuvchi ma'lumotlarini tekshirish
        if chat_id in user_settings:
            current_subject = user_settings[chat_id].get('subject', 'Tanlanmagan')
            current_sections = user_settings[chat_id].get('sections', {})
            
            current_info = f"🔧 *Joriy sozlamalar:*\n📚 Fan: {current_subject}\n📖 Bo'limlar: {len(current_sections)} ta\n\n"
        else:
            current_info = "🔧 *Joriy sozlamalar:*\n📚 Fan: Tanlanmagan\n📖 Bo'limlar: 0 ta\n\n"
        
        # Fanlar ro'yxatini yaratish
        markup = types.InlineKeyboardMarkup(row_width=2)
        
        for key, subject in SUBJECTS.items():
            btn = types.InlineKeyboardButton(
                f"📚 {subject['name']}", 
                callback_data=f'set_subject_{key}'
            )
            markup.add(btn)
        
        setting_text = current_info + "📚 *Milliy sertifikat fanalaridan birini tanlang:*"
        
        bot.reply_to(message, setting_text, parse_mode='Markdown', reply_markup=markup)

    # Text message handler
    @bot.message_handler(content_types=['text'])
    def handle_text(message):
        """Text message handler"""
        chat_id = message.chat.id
        text = message.text.strip()
        
        # Check if user is waiting for ball input
        if chat_id in user_data and user_data[chat_id].get('waiting_for_ball'):
            handle_ball_input(message)
        # Check if user is waiting for section questions
        elif chat_id in user_settings and user_settings[chat_id].get('waiting_for_questions'):
            handle_section_questions(message)
        else:
            # Default response for text messages
            bot.reply_to(message, "📝 Iltimos, Excel faylni yuboring yoki /help buyrug'ini ishlating.")

    def handle_section_questions(message):
        """Bo'lim savollarini qayta ishlash"""
        chat_id = message.chat.id
        text = message.text.strip()
        
        if chat_id not in user_settings:
            bot.reply_to(message, "❌ Xatolik: Sozlamalar topilmadi. /setting buyrug'ini ishlating.")
            return
        
        current_section = user_settings[chat_id]['current_section_name']
        subject_key = user_settings[chat_id]['subject_key']
        
        try:
            # Savol raqamlarini ajratish
            question_numbers = []
            for part in text.split(','):
                part = part.strip()
                if part.isdigit():
                    question_numbers.append(int(part))
            
            if not question_numbers:
                bot.reply_to(message, "❌ Xatolik: Savol raqamlarini to'g'ri kiriting!\n\n💡 *Misol:* 1, 3, 5, 7, 9", parse_mode='Markdown')
                return
            
            # Savollarni saqlash
            user_settings[chat_id]['sections'][current_section] = question_numbers
            user_settings[chat_id]['waiting_for_questions'] = False
            
            # Keyingi bo'limga o'tish
            sections = SUBJECTS[subject_key]['sections']
            current_index = user_settings[chat_id]['current_section']
            next_index = current_index + 1
            
            if next_index < len(sections):
                # Keyingi bo'limni ko'rsatish
                next_section = sections[next_index]
                markup = types.InlineKeyboardMarkup()
                next_btn = types.InlineKeyboardButton(
                    f"📖 {next_section}", 
                    callback_data=f'set_section_{subject_key}_{next_index}'
                )
                skip_btn = types.InlineKeyboardButton("⏭️ O'tkazib yuborish", callback_data=f'skip_section_{subject_key}_{next_index}')
                markup.add(next_btn)
                markup.add(skip_btn)
                
                success_text = f"✅ *{current_section} bo'limi saqlandi!*\n\n"
                success_text += f"📊 Savollar: {len(question_numbers)} ta\n"
                success_text += f"📝 Raqamlar: {', '.join(map(str, question_numbers))}\n\n"
                success_text += f"📖 *Keyingi bo'lim:* {next_section}"
                
                bot.reply_to(message, success_text, parse_mode='Markdown', reply_markup=markup)
            else:
                # Barcha bo'limlar tugadi
                total_sections = len(user_settings[chat_id]['sections'])
                total_questions = sum(len(questions) for questions in user_settings[chat_id]['sections'].values())
                
                success_text = f"✅ *Sozlash muvaffaqiyatli tugadi!*\n\n"
                success_text += f"📚 Fan: {user_settings[chat_id]['subject']}\n"
                success_text += f"📖 Bo'limlar: {total_sections} ta\n"
                success_text += f"📊 Jami savollar: {total_questions} ta\n\n"
                success_text += "📊 *Endi natijalar matritsasini yuborishingiz mumkin!*"
                
                bot.reply_to(message, success_text, parse_mode='Markdown')
                
        except Exception as e:
            bot.reply_to(message, f"❌ Xatolik: {str(e)}\n\n💡 Iltimos, qaytadan urinib ko'ring.")

    # Start the bot
    print("🚀 Bot ishga tushirilmoqda...")
    bot.polling(none_stop=True)

if __name__ == '__main__':
    main()
